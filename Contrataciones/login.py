import tkinter as tk
from tkinter import ttk, messagebox
import pymysql
from conexion import *
import pandas as pd
from tkcalendar import DateEntry
import mysql.connector
from tkinter import messagebox
from tkinter import filedialog
import pandas as pd
import os
import math
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import re 
from datetime import datetime



# Configuración de la base de datos
class DataBase:
    def __init__(self):
        self.connection = pymysql.connect(
            host="localhost",
            user="root",
            password="",
            db="contrataciones"
        )
        self.cursor = self.connection.cursor()

    def __del__(self):
        self.cursor.close()
        self.connection.close()

# Funciones de la aplicación
def menu_pantalla():
    global pantalla
    pantalla = tk.Tk()
    pantalla.geometry("300x440")
    pantalla.title("CENIGRAF")

    image = tk.PhotoImage(file="logo.gif")
    image = image.subsample(2, 2)
    label = tk.Label(image=image)
    label.pack()

    tk.Label(text="Acceso al sistema", bg="green", fg="white", width="300", height="3", font=("Calibri", 15)).pack()
    tk.Label(text="").pack()

    tk.Button(text="Iniciar Sesion", height="3", width="30", command=inicio_sesion).pack()
    tk.Label(text="").pack()

    tk.Button(text="Registrar", height="3", width="30", command=registrar).pack()

    pantalla.mainloop()

def regresar_a_pantalla_principal(window):
    window.destroy()

def crear_ventana(pantalla_principal, title, size):
    ventana = tk.Toplevel(pantalla_principal)
    ventana.geometry(size)
    ventana.title(title)

    # Agregar el botón de regresar
    tk.Button(
        ventana, text="Regresar", bg="blue", fg="white", command=lambda: regresar_a_pantalla_principal(ventana)
    ).pack(side="top", anchor="nw", padx=5, pady=5)

    return ventana

def inicio_sesion():
    global pantalla1
    pantalla1 = crear_ventana(pantalla, "Inicio de sesión", "400x310")

    tk.Label(pantalla1, text="Ingrese su usuario y contraseña", bg="green", fg="white", width="300", height="3", font=("Calibri", 15)).pack()
    tk.Label(pantalla1, text="").pack()

    global nombreusuario_verify
    global contrasenausaurio_verify

    nombreusuario_verify = tk.StringVar()
    contrasenausaurio_verify = tk.StringVar()

    global nombre_usuario_entry
    global contrasena_usaurio_entry

    tk.Label(pantalla1, text="Usuario").pack()
    nombre_usuario_entry = tk.Entry(pantalla1, textvariable=nombreusuario_verify)
    nombre_usuario_entry.pack()
    tk.Label(pantalla1).pack()

    tk.Label(pantalla1, text="Contraseña").pack()
    contrasena_usaurio_entry = tk.Entry(pantalla1, show="*", textvariable=contrasenausaurio_verify)
    contrasena_usaurio_entry.pack()
    tk.Label(pantalla1).pack()

    tk.Button(pantalla1, text="Iniciar Sesión", command=validacion_datos).pack()

def registrar():
    global pantalla2
    pantalla2 = crear_ventana(pantalla, "Registrar", "400x380")

    global nombreusuario_entry
    global correo_entry
    global contrasenausaurio_entry

    tk.Label(pantalla2, text="Ingrese un usuario y contraseña,\n para el registro del sistema", bg="green", fg="white", width="300", height="3", font=("Calibri", 15)).pack()
    tk.Label(pantalla2, text="").pack()

    tk.Label(pantalla2, text="Número de documento").pack()
    nombreusuario_entry = tk.Entry(pantalla2)
    nombreusuario_entry.pack()
    tk.Label(pantalla2).pack()

    tk.Label(pantalla2, text="Correo electrónico").pack()
    correo_entry = tk.Entry(pantalla2)
    correo_entry.pack()
    tk.Label(pantalla2).pack()

    tk.Label(pantalla2, text="Contraseña").pack()
    contrasenausaurio_entry = tk.Entry(pantalla2, show="*")
    contrasenausaurio_entry.pack()
    tk.Label(pantalla2).pack()

    tk.Button(pantalla2, text="Registrar", command=inserta_datos).pack()

def inserta_datos():
    # Conectar a la base de datos
    bd = pymysql.connect(
        host="localhost",
        user="root",
        password="",
        db="contrataciones"
    )
    fcursor = bd.cursor()

    # Preparar la consulta SQL con placeholders
    sql = "INSERT INTO usuarios(numeroCedulaCliente, correoCli, contraseñaCli) VALUES (%s, %s, %s)"
    datos = (nombreusuario_entry.get(), correo_entry.get(), contrasenausaurio_entry.get())

    try:
        fcursor.execute(sql, datos)
        bd.commit()
        messagebox.showinfo(message="Registro exitoso", title="Aviso")
    except pymysql.MySQLError as e:
        bd.rollback()
        messagebox.showinfo(message=f"No Registrado. Error: {e}", title="Aviso")
    finally:
        fcursor.close()
        bd.close()

def validacion_datos():
    global usuario_iniciado

    # Conectar a la base de datos
    bd = pymysql.connect(
        host="localhost",
        user="root",
        password="",
        db="contrataciones"
    )
    fcursor = bd.cursor()

    usuario = nombreusuario_verify.get()
    contrasena = contrasenausaurio_verify.get()

    # Validar campos vacíos
    if not usuario or not contrasena:
        messagebox.showinfo(title="Error", message="Por favor, ingrese ambos campos")
        return

    # Ejecutar consulta SQL con placeholders
    sql = "SELECT contraseñaCli FROM usuarios WHERE correoCli=%s AND contraseñaCli=%s"
    fcursor.execute(sql, (usuario, contrasena))

    if fcursor.fetchone():
        usuario_iniciado = True
        pantalla_bienvenida()
    else:
        messagebox.showinfo(title="Inicio de sesión incorrecto", message="Usuario y contraseña incorrectos")

    fcursor.close()
    bd.close()

estados = {
    "Pantalla ARL": 2,
    "Pantalla EPS": 2,
    "Pantalla Bancos": 1,
    "Pantalla Ciudad": 1,
    "Pantalla Cargo": 1,
    "Pantalla Tipo de contrato": 1,
    "Pantalla Jefe": 1,
    "Pantalla Dependencia": 1,
    "Pantalla Contrato": 1,
    "Pantalla Contratistas": 1,
    "Pantalla Departamento": 2,
    "Pantalla RH": 2,
    "Pantalla Género": 2,
    "Pantalla Tipo de documento": 2,
    "Pantalla Tipo de Cuenta": 1,
    "Pantalla Registro Usuarios": 1,
    "Pantalla PAA":1
}

def pantalla_bienvenida():
    if not usuario_iniciado:
        messagebox.showinfo(title="Acceso denegado", message="Debe iniciar sesión para acceder a esta pantalla")
        return

    # Crear una nueva ventana
    bienvenida = crear_ventana(pantalla1, "Bienvenida", "800x200")

    tk.Label(bienvenida, text="Pantallas de Funcionalidad", font=("Calibri", 15)).pack(pady=20)

    # Crear un marco para los botones
    frame_botones = tk.Frame(bienvenida)
    frame_botones.pack(padx=10, pady=10)

    # Lista de textos para los botones
    boton_textos = [
        "Pantalla ARL",
        "Pantalla EPS",
        "Pantalla Bancos", 
        "Pantalla Ciudad",
        "Pantalla Cargo", 
        "Pantalla Tipo de contrato", 
        "Pantalla Jefe",
        "Pantalla Dependencia", 
        "Pantalla Contrato", 
        "Pantalla Contratistas",
        "Pantalla Departamento",
        "Pantalla RH",
        "Pantalla Género",
        "Pantalla Tipo de documento", 
        "Pantalla Tipo de Cuenta", 
        "Pantalla Registro Usuarios",
        "Pantalla PAA"

    ]

    # Añadir los botones al marco utilizando grid, sin dejar espacios vacíos
    row = 0
    column = 0
    for texto in boton_textos:
        # Verificar si el botón debe ser visible
        if estados.get(texto) == 1:
            boton = tk.Button(frame_botones, text=texto, bg="white", fg="black",
                            command=lambda t=texto: abrir_ventana_funcionalidad(t))
            boton.grid(row=row, column=column, padx=5, pady=5, sticky="ew")
            
            # Actualizar la columna, y si llega a 5, reiniciarla y pasar a la siguiente fila
            column += 1
            if column == 5:
                column = 0
                row += 1

    # Botón para cerrar sesión
    tk.Button(bienvenida, text="Cerrar Sesión", bg="red", fg="white", command=cerrar_sesion).pack(pady=10)

def abrir_ventana_funcionalidad(titulo):
    ventana = crear_ventana(pantalla1, titulo, "500x400")

    # Aquí puedes definir el CRUD para cada funcionalidad
    if titulo == "Pantalla ARL":
        mostrar_arl(ventana)
    elif titulo == "Pantalla EPS":
        mostrar_eps(ventana)
    elif titulo == "Pantalla Bancos":
        mostrar_bancos(ventana)
    elif titulo == "Pantalla Ciudad":
        mostrar_ciudad(ventana)
    elif titulo == "Pantalla Cargo":
        mostrar_cargo(ventana)
    elif titulo == "Pantalla Tipo de contrato":
        mostrar_tipodecontrato(ventana)
    elif titulo == "Pantalla Dependencia":
        mostrar_dependencia(ventana)
    elif titulo == "Pantalla Departamento":
        mostrar_departamento(ventana)
    elif titulo == "Pantalla RH":
        mostrar_rh(ventana)
    elif titulo == "Pantalla Género":
        mostrar_genero(ventana)
    elif titulo == "Pantalla Tipo de documento":
        mostrar_tipodedocumento(ventana) 
    elif titulo == "Pantalla Tipo de Cuenta":
        mostrar_tipo_cuenta(ventana)       
    elif titulo == "Pantalla Contrato":
        mostrar_contrato(ventana)   
    elif titulo == "Pantalla Contratistas":
        mostrar_Contratistas(ventana)
    elif titulo == "Pantalla Jefe":
        mostrar_jefes(ventana)
    elif titulo == "Pantalla Registro Usuarios":
        mostrar_usuario(ventana)
    elif titulo == "Pantalla PAA":
        mostrar_PAA(ventana)
    # Agrega más condiciones para otras pantallas aquí
def exportar_a_excel_arl():
    try:
        sql = "SELECT id, nombreARL FROM arl"
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        if filas:
            df = pd.DataFrame(filas, columns=['ID', 'Nombre'])

            filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                    filetypes=[("Excel files", "*.xlsx"), 
                                                               ("All files", "*.*")])
            
            if filepath:
                df.to_excel(filepath, index=False)
                messagebox.showinfo("Exportación Exitosa", f"Datos exportados a {filepath}")
                os.startfile(filepath)
            else:
                messagebox.showwarning("Cancelado", "Exportación cancelada por el usuario")
        else:
            messagebox.showwarning("Sin Datos", "No hay datos para exportar")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar los datos: {e}")

def importar_desde_excel_arl():
    try:
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        
        if filepath:
            df = pd.read_excel(filepath)
            
            if 'ID' in df.columns and 'Nombre' in df.columns:
                for _, row in df.iterrows():
                    id_arl = row['ID']
                    nombre_arl = row['Nombre']
                    sql = "INSERT INTO arl (id, nombreARL) VALUES (%s, %s)"
                    db.cursor.execute(sql, (id_arl, nombre_arl))
                db.connection.commit()
                messagebox.showinfo("Importación Exitosa", "Datos importados correctamente desde el archivo Excel")
                llenar_tabla()  # Refresca la tabla con los nuevos datos importados
            else:
                messagebox.showerror("Error", "El archivo Excel no tiene las columnas correctas ('ID' y 'Nombre')")
        else:
            messagebox.showwarning("Cancelado", "Importación cancelada por el usuario")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo importar los datos: {e}")

def mostrar_arl(ventana):
    global db
    db = DataBase()
    modificar = False
    dni = tk.StringVar()
    nombre = tk.StringVar()

    def seleccionar(event):
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvEstudiantes.item(id, "values")
            dni.set(valores[0])
            nombre.set(valores[1])

    marco = tk.LabelFrame(ventana, text="Formulario ARL")
    marco.place(x=50, y=50, width=500, height=400)

    tk.Label(marco, text="DNI").grid(column=0, row=0, padx=5, pady=5)
    txtDni = tk.Entry(marco, textvariable=dni)
    txtDni.grid(column=1, row=0)

    tk.Label(marco, text="Nombre").grid(column=0, row=1, padx=5, pady=5)
    txtNombre = tk.Entry(marco, textvariable=nombre)
    txtNombre.grid(column=1, row=1)

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=2, columnspan=4)

    tvEstudiantes = ttk.Treeview(marco, selectmode='none')
    tvEstudiantes["columns"] = ("DNI", "Nombre")
    tvEstudiantes.column("#0", width=0, stretch='no')
    tvEstudiantes.column("DNI", width=150, anchor='center')
    tvEstudiantes.column("Nombre", width=150, anchor='center')
    tvEstudiantes.heading("#0", text="")
    tvEstudiantes.heading("DNI", text="DNI", anchor='center')
    tvEstudiantes.heading("Nombre", text="Nombre", anchor='center')
    tvEstudiantes.grid(column=0, row=3, columnspan=4, padx=5)
    tvEstudiantes.bind("<<TreeviewSelect>>", seleccionar)

    btnEliminar = tk.Button(marco, text="Eliminar", command=lambda: eliminar())
    btnEliminar.grid(column=1, row=4)
    btnNuevo = tk.Button(marco, text="Guardar", command=lambda: nuevo())
    btnNuevo.grid(column=2, row=4)
    btnModificar = tk.Button(marco, text="Seleccionar", command=lambda: actualizar())
    btnModificar.grid(column=3, row=4)

    # Botón de exportar a Excel
    btnExportar = tk.Button(marco, text="Exportar a Excel", command=exportar_a_excel_arl)
    btnExportar.grid(column=2, row=5, pady=10)

    # Botón de importar desde Excel
    btnImportar = tk.Button(marco, text="Importar desde Excel", command=importar_desde_excel_arl)
    btnImportar.grid(column=3, row=5, pady=10)

    # Agregar los botones para cerrar y minimizar
    btnCerrar = tk.Button(marco, text="X", command=ventana.destroy, bg="red", fg="white")
    btnCerrar.place(x=470, y=10, width=20, height=20)
    
    btnMinimizar = tk.Button(marco, text="-", command=lambda: ventana.iconify(), bg="yellow", fg="black")
    btnMinimizar.place(x=450, y=10, width=20, height=20)

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tvEstudiantes.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Seleccionar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tvEstudiantes.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def validar():
        return len(dni.get()) > 0 and len(nombre.get()) > 0

    def limpiar():
        dni.set("")
        nombre.set("")

    def vaciar_tabla():
        filas = tvEstudiantes.get_children()
        for fila in filas:
            tvEstudiantes.delete(fila)

    def llenar_tabla():
        vaciar_tabla()
        sql = "SELECT id, nombreARL FROM arl"
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        for fila in filas:
            id = fila[0]
            tvEstudiantes.insert("", 'end', id, text=id, values=(fila[0], fila[1]))

    def eliminar():
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM arl WHERE id=%s"
                db.cursor.execute(sql, (id,))
                db.connection.commit()
                tvEstudiantes.delete(id)
                lblMensaje.config(text="Se ha eliminado el registro correctamente")
                limpiar()
                llenar_tabla()
            else:
                lblMensaje.config(text="Seleccione un registro para eliminar")

    def nuevo():
        if not modificar:
            if validar():
                val = (dni.get(), nombre.get())
                sql = "INSERT INTO arl (id, nombreARL) VALUES (%s, %s)"
                db.cursor.execute(sql, val)
                db.connection.commit()
                lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                llenar_tabla()
                limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarFalse()

    def actualizar():
        if modificar:
            if validar():
                seleccion = tvEstudiantes.selection()
                if seleccion:
                    id = seleccion[0]
                    val = (dni.get(), nombre.get())
                    sql = "UPDATE arl SET id=%s, nombreARL=%s WHERE id=%s"
                    db.cursor.execute(sql, val + (id,))
                    db.connection.commit()
                    lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                    llenar_tabla()
                    limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarTrue()

    llenar_tabla()

def exportar_a_excel_tipodecontrato():
    try:
        sql = "SELECT id, nombreTipoContrato FROM tipodecontrato"
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        if filas:
            df = pd.DataFrame(filas, columns=['ID Tipo', 'Nombre Tipo'])
            
            filepath = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                    filetypes=[("Excel files", "*.xlsx"),
                                                               ("All files", "*.*")])
            
            if filepath:
                df.to_excel(filepath, index=False)
                messagebox.showinfo("Exportación Exitosa", f"Datos exportados a {filepath}")
                os.startfile(filepath)
            else:
                messagebox.showwarning("Cancelado", "Exportación cancelada por el usuario")
        else:
            messagebox.showwarning("Sin Datos", "No hay datos para exportar")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar los datos: {e}")

def importar_desde_excel_tipodecontrato():
    try:
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        
        if filepath:
            df = pd.read_excel(filepath)
            
            if 'ID Tipo' in df.columns and 'Nombre Tipo' in df.columns:
                db.cursor.execute("DELETE FROM tipodecontrato") 
                for _, row in df.iterrows():
                    id_tipo = row['ID Tipo']
                    nombre_tipo = row['Nombre Tipo']
                    sql = "INSERT INTO tipodecontrato (id, nombreTipoContrato) VALUES (?, ?)"
                    db.cursor.execute(sql, (id_tipo, nombre_tipo))
                db.connection.commit()
                messagebox.showinfo("Importación Exitosa", "Datos importados correctamente desde el archivo Excel")
                llenar_tabla()  # Refresca la tabla con los nuevos datos importados
            else:
                messagebox.showerror("Error", "El archivo Excel no tiene las columnas correctas ('ID Tipo', 'Nombre Tipo')")
        else:
            messagebox.showwarning("Cancelado", "Importación cancelada por el usuario")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo importar los datos: {e}")

def mostrar_tipodecontrato(ventana):
    global db, page_number, records_per_page, total_records
    db = DataBase()
    modificar = False
    id_tipo = tk.StringVar()
    nombre_tipo = tk.StringVar()
    filtro_nombre_tipo = tk.StringVar()  # Variable para el filtro del combobox

    # Variables de paginación
    page_number = 1
    records_per_page = 10
    total_records = 0

    def seleccionar(event):
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvEstudiantes.item(id, "values")
            id_tipo.set(valores[0])
            nombre_tipo.set(valores[1])
            modificarTrue()

    marco = tk.LabelFrame(ventana, text="Formulario Tipo de Contrato")
    marco.place(x=50, y=50, width=700, height=600)

    tk.Label(marco, text="ID Tipo").grid(column=0, row=0, padx=5, pady=5)
    txtIdTipo = tk.Entry(marco, textvariable=id_tipo)
    txtIdTipo.grid(column=1, row=0)

    tk.Label(marco, text="Nombre Tipo").grid(column=0, row=1, padx=5, pady=5)
    txtNombreTipo = tk.Entry(marco, textvariable=nombre_tipo)
    txtNombreTipo.grid(column=1, row=1)

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=2, columnspan=4)

    tvEstudiantes = ttk.Treeview(marco, selectmode='none')
    tvEstudiantes["columns"] = ("ID Tipo", "Nombre Tipo")
    tvEstudiantes.column("#0", width=0, stretch='no')
    tvEstudiantes.column("ID Tipo", width=150, anchor='center')
    tvEstudiantes.column("Nombre Tipo", width=150, anchor='center')
    tvEstudiantes.heading("#0", text="")
    tvEstudiantes.heading("ID Tipo", text="ID Tipo", anchor='center')
    tvEstudiantes.heading("Nombre Tipo", text="Nombre Tipo", anchor='center')
    tvEstudiantes.grid(column=0, row=3, columnspan=4, padx=5, pady=10)
    tvEstudiantes.bind("<<TreeviewSelect>>", seleccionar)

    # Combobox para el filtro por Nombre Tipo
    tk.Label(marco, text="Filtrar por Nombre Tipo").grid(column=0, row=4, padx=5, pady=5)
    comboFiltroNombreTipo = ttk.Combobox(marco, textvariable=filtro_nombre_tipo)
    comboFiltroNombreTipo.grid(column=1, row=4, padx=5, pady=5)

    def cargar_tipos():
        sql = "SELECT nombreTipoContrato FROM tipodecontrato"
        db.cursor.execute(sql)
        tipos = [fila[0] for fila in db.cursor.fetchall()]
        comboFiltroNombreTipo["values"] = tipos

    cargar_tipos()

    def limpiar_campos():
        """Función para limpiar todos los campos después de aplicar el filtro."""
        id_tipo.set("")
        nombre_tipo.set("")
        filtro_nombre_tipo.set("")  # Limpiar el combobox

    def aplicar_filtro(event=None):
        llenar_tabla(filtro=filtro_nombre_tipo.get())
        limpiar_campos()  # Limpiar los campos después de aplicar el filtro

    # Vincular el evento de selección del combobox para aplicar el filtro automáticamente
    comboFiltroNombreTipo.bind("<<ComboboxSelected>>", aplicar_filtro)

    # Botón para mostrar todos los registros
    btnEliminarFiltro = tk.Button(marco, text="Mostrar Consultas", command=lambda: llenar_tabla(filtro=""))
    btnEliminarFiltro.grid(column=3, row=4, padx=5, pady=5)

    btnEliminar = tk.Button(marco, text="Eliminar", command=lambda: eliminar())
    btnEliminar.grid(column=1, row=5, padx=5, pady=10)
    btnNuevo = tk.Button(marco, text="Guardar", command=lambda: nuevo())
    btnNuevo.grid(column=2, row=5, padx=5, pady=10)
    btnModificar = tk.Button(marco, text="Seleccionar", command=lambda: actualizar())
    btnModificar.grid(column=3, row=5, padx=5, pady=10)

    btnExportar = tk.Button(marco, text="Exportar a Excel", command=exportar_a_excel_tipodecontrato)
    btnExportar.grid(column=2, row=6, padx=5, pady=10)

    btnImportar = tk.Button(marco, text="Importar desde Excel", command=importar_desde_excel_tipodecontrato)
    btnImportar.grid(column=3, row=6, padx=5, pady=10)

    btnCerrar = tk.Button(marco, text="X", command=ventana.destroy, bg="red", fg="white")
    btnCerrar.place(x=660, y=10, width=30, height=30)
    
    btnMinimizar = tk.Button(marco, text="-", command=lambda: ventana.iconify(), bg="yellow", fg="black")
    btnMinimizar.place(x=620, y=10, width=30, height=30)

    # Paginación
    btnPrev = tk.Button(marco, text="<< Anterior", command=lambda: cambiar_pagina(-1))
    btnPrev.grid(column=0, row=7, padx=5, pady=10)

    btnNext = tk.Button(marco, text="Siguiente >>", command=lambda: cambiar_pagina(1))
    btnNext.grid(column=3, row=7, padx=5, pady=10)

    def cambiar_pagina(direccion):
        global page_number
        page_number += direccion
        llenar_tabla(filtro=filtro_nombre_tipo.get())

    def actualizar_botones_paginacion():
        global total_records
        btnPrev.config(state='normal' if page_number > 1 else 'disabled')
        btnNext.config(state='normal' if page_number * records_per_page < total_records else 'disabled')

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tvEstudiantes.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Seleccionar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tvEstudiantes.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def validar():
        return len(id_tipo.get()) > 0 and len(nombre_tipo.get()) > 0

    def limpiar():
        id_tipo.set("")
        nombre_tipo.set("")

    def vaciar_tabla():
        filas = tvEstudiantes.get_children()
        for fila in filas:
            tvEstudiantes.delete(fila)

    def llenar_tabla(filtro=""):
        vaciar_tabla()
        global total_records

        offset = (page_number - 1) * records_per_page

        if filtro:
            # Consulta con filtro
            sql_count = "SELECT COUNT(*) FROM tipodecontrato WHERE nombreTipoContrato LIKE %s"
            db.cursor.execute(sql_count, ('%' + filtro + '%',))
            total_records = db.cursor.fetchone()[0]

            sql = f"SELECT id, nombreTipoContrato FROM tipodecontrato WHERE nombreTipoContrato LIKE %s LIMIT {records_per_page} OFFSET {offset}"
            db.cursor.execute(sql, ('%' + filtro + '%',))
        else:
            # Consulta sin filtro
            sql_count = "SELECT COUNT(*) FROM tipodecontrato"
            db.cursor.execute(sql_count)
            total_records = db.cursor.fetchone()[0]

            sql = f"SELECT id, nombreTipoContrato FROM tipodecontrato LIMIT {records_per_page} OFFSET {offset}"
            db.cursor.execute(sql)

        filas = db.cursor.fetchall()

        for fila in filas:
            id = fila[0]
            tvEstudiantes.insert("", 'end', id, text=id, values=(fila[0], fila[1]))

        actualizar_botones_paginacion()

    def nuevo():
        if not modificar:
            if validar():
                val = (id_tipo.get(), nombre_tipo.get())
                sql = "INSERT INTO tipodecontrato (id, nombreTipoContrato) VALUES (%s, %s)"
                db.cursor.execute(sql, val)
                db.connection.commit()
                lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                llenar_tabla()
                limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarFalse()

    def eliminar():
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM tipodecontrato WHERE id=%s"
                db.cursor.execute(sql, (id,))
                db.connection.commit()
                tvEstudiantes.delete(id)
                lblMensaje.config(text="Se ha eliminado el registro correctamente")
                limpiar()
                llenar_tabla()

    def actualizar():
        if modificar:
            if validar():
                val = (nombre_tipo.get(), id_tipo.get())
                sql = "UPDATE tipodecontrato SET nombreTipoContrato=%s WHERE id=%s"
                db.cursor.execute(sql, val)
                db.connection.commit()
                lblMensaje.config(text="Se ha modificado el registro con éxito", fg="green")
                llenar_tabla()
                limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarTrue()

    llenar_tabla()

def mostrar_Contratistas(ventana):
    # Configuración de la conexión a la base de datos
    db = mysql.connector.connect(host="localhost", user="root", password="", database="contrataciones")
    cursor = db.cursor()

    # Crear un nuevo marco dentro del Toplevel para usar grid
    frame_principal = tk.Frame(ventana)
    frame_principal.pack(fill='both', expand=True)

    # Configuración del marco principal
    marco = tk.LabelFrame(frame_principal, text="Formulario Contratistas")
    marco.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

    # Hacer que el marco sea redimensionable
    frame_principal.grid_rowconfigure(0, weight=1)
    frame_principal.grid_columnconfigure(0, weight=1)
    marco.grid_rowconfigure(11, weight=1)  # Permitir que la fila de la tabla se expanda
    marco.grid_columnconfigure(5, weight=1)  # Permitir que las columnas de widgets se expandan

    # Variables de paginación
    registros_por_pagina = 10
    pagina_actual = 0
    total_paginas = 0

    # Función para obtener opciones para los comboboxes
    def obtener_opciones(query):
        cursor.execute(query)
        return [(fila[0], fila[1]) for fila in cursor.fetchall()]

    # Consultas para llenar los comboboxes
    opciones_tipo_documento = obtener_opciones("SELECT id, numeroCedula FROM tipodedocumento")
    opciones_departamento = obtener_opciones("SELECT id, nombreDepartamento FROM departamentos")
    opciones_ciudad = obtener_opciones("SELECT id, nombreCiudad FROM ciudad")
    opciones_genero = obtener_opciones("SELECT id, nombreGenero FROM genero")
    opciones_rh = obtener_opciones("SELECT id, tipoRh FROM rh")
    opciones_arl = obtener_opciones("SELECT id, nombreARL FROM arl")
    opciones_eps = obtener_opciones("SELECT id, nombreEPS FROM eps")
    opciones_banco = obtener_opciones("SELECT id, nombreBanco FROM banco")
    opciones_estado = obtener_opciones("SELECT id, tipoEstado FROM estado")
    opciones_usuario_creador = obtener_opciones("SELECT id, numeroCedulaCliente FROM usuarios")

    # Variables de los campos de la tabla clientes
    id_cliente = tk.StringVar()
    numero_documento = tk.StringVar()
    primerNombre = tk.StringVar()
    segundoNombre = tk.StringVar()
    primerApellido = tk.StringVar()
    segundoApellido = tk.StringVar()
    tipoDocumento = tk.StringVar()
    departamentoExpedicion = tk.StringVar()
    ciudadExpedicion = tk.StringVar()
    fechaExpedicion = tk.StringVar()
    genero = tk.StringVar()
    fechaNacimento = tk.StringVar()
    Rh = tk.StringVar()
    dirrecion = tk.StringVar()
    correo = tk.StringVar()
    correoAdicional = tk.StringVar()
    celular = tk.StringVar()
    telefono = tk.StringVar()
    ARL = tk.StringVar()
    EPS = tk.StringVar()
    PAA = tk.StringVar()
    idBanco = tk.StringVar()
    numeroCuenta = tk.StringVar()
    CDP = tk.StringVar()
    fechaRegistro = tk.StringVar()
    ultimoAcceso = tk.StringVar()
    Estado = tk.StringVar()
    usuarioCreador = tk.StringVar()

    def validar_correo(correo):
        patron_correo = r"^[\w\.-]+@[\w\.-]+\.\w+$"
        return re.match(patron_correo, correo) is not None

    # Función para validar que todos los campos requeridos estén llenos
    def validar_campos():
        campos_faltantes = []
        if not numero_documento.get(): campos_faltantes.append("Número Documento")
        if not primerNombre.get(): campos_faltantes.append("Primer Nombre")
        if not segundoNombre.get(): campos_faltantes.append("Segundo Nombre")
        if not primerApellido.get(): campos_faltantes.append("Primer Apellido")
        if not segundoApellido.get(): campos_faltantes.append("Segundo Apellido")
        if not tipoDocumento.get(): campos_faltantes.append("Tipo Documento")
        if not departamentoExpedicion.get(): campos_faltantes.append("Departamento Expedición")
        if not ciudadExpedicion.get(): campos_faltantes.append("Ciudad Expedición")
        if not fechaExpedicion.get(): campos_faltantes.append("Fecha Expedición")
        if not genero.get(): campos_faltantes.append("Género")
        if not fechaNacimento.get(): campos_faltantes.append("Fecha Nacimiento")
        if not Rh.get(): campos_faltantes.append("RH")
        if not dirrecion.get(): campos_faltantes.append("Dirección")
        if not correo.get(): campos_faltantes.append("Correo")
        if not correoAdicional.get(): campos_faltantes.append("Correo Adicional")
        if not celular.get(): campos_faltantes.append("Celular")
        if not telefono.get(): campos_faltantes.append("Teléfono")
        if not ARL.get(): campos_faltantes.append("ARL")
        if not EPS.get(): campos_faltantes.append("EPS")
        if not PAA.get(): campos_faltantes.append("PAA")
        if not idBanco.get(): campos_faltantes.append("Banco")
        if not numeroCuenta.get(): campos_faltantes.append("Número Cuenta")
        if not CDP.get(): campos_faltantes.append("CDP")
        if not fechaRegistro.get(): campos_faltantes.append("Fecha Registro")
        if not ultimoAcceso.get(): campos_faltantes.append("Último Acceso")
        if not Estado.get(): campos_faltantes.append("Estado")
        if not usuarioCreador.get(): campos_faltantes.append("Usuario Creador")
        if not validar_correo(correo.get()): campos_faltantes.append("Correo válido")

        if campos_faltantes:
            lblMensaje.config(text=f"Faltan campos: {', '.join(campos_faltantes)}", fg="red", font=("Arial", 12, "bold"))
            lblMensaje.grid(columnspan=6, sticky="ew")
            ajustar_texto_mensaje()
            return False
        return True

    def seleccionar(event):
        seleccion = tvClientes.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvClientes.item(id, "values")

            # Asegurarnos de que el número de columnas coincida
            if len(valores) == len(tvClientes["columns"]):
                id_cliente.set(valores[0])
                numero_documento.set(valores[0])
                primerNombre.set(valores[1])
                segundoNombre.set(valores[2])
                primerApellido.set(valores[3])
                segundoApellido.set(valores[4])
                tipoDocumento.set(valores[5])
                departamentoExpedicion.set(valores[6])
                ciudadExpedicion.set(valores[7])
                fechaExpedicion.set(valores[8])
                genero.set(valores[9])
                fechaNacimento.set(valores[10])
                Rh.set(valores[11])
                dirrecion.set(valores[12])
                correo.set(valores[13])
                correoAdicional.set(valores[14])
                celular.set(valores[15])
                telefono.set(valores[16])
                ARL.set(valores[17])
                EPS.set(valores[18])
                PAA.set(valores[19])
                idBanco.set(valores[20])
                numeroCuenta.set(valores[21])
                CDP.set(valores[22])
                fechaRegistro.set(valores[23])
                ultimoAcceso.set(valores[24])
                Estado.set(valores[25])
                usuarioCreador.set(valores[26])

                # Bloquear el campo de Número de Documento al editar
                numero_documento_entry.config(state="readonly")
            else:
                lblMensaje.config(text="Error: Los datos seleccionados no tienen suficientes columnas.", fg="red", font=("Arial", 12, "bold"))
                ajustar_texto_mensaje()

    def limpiar_campos():
        id_cliente.set("")
        numero_documento.set("")
        primerNombre.set("")
        segundoNombre.set("")
        primerApellido.set("")
        segundoApellido.set("")
        tipoDocumento.set("")
        departamentoExpedicion.set("")
        ciudadExpedicion.set("")
        fechaExpedicion.set("")
        genero.set("")
        fechaNacimento.set("")
        Rh.set("")
        dirrecion.set("")
        correo.set("")
        correoAdicional.set("")
        celular.set("")
        telefono.set("")
        ARL.set("")
        EPS.set("")
        PAA.set("")
        idBanco.set("")
        numeroCuenta.set("")
        CDP.set("")
        fechaRegistro.set("")
        ultimoAcceso.set("")
        Estado.set("")
        usuarioCreador.set("")

        # Desbloquear el campo de Número de Documento
        numero_documento_entry.config(state="normal")

    # Función para filtrar las opciones en los comboboxes
    def filtrar_opciones(opciones, texto):
        texto_filtrado = texto.lower()
        return [txt for id, txt in opciones if texto_filtrado in txt.lower()]

    def actualizar_combobox(event, opciones, combobox):
        texto_actual = combobox.get()
        opciones_filtradas = filtrar_opciones(opciones, texto_actual)
        combobox['values'] = opciones_filtradas
        combobox.icursor(len(texto_actual))  # Mover el cursor al final del texto
        combobox.event_generate('<Down>')  # Mostrar lista desplegable

    # Configuración del marco de formulario
    numero_documento_entry = tk.Entry(marco, textvariable=numero_documento, width=20, validate="key", validatecommand=(ventana.register(lambda s: s.isdigit()), "%P"))
    
    tipoDocumento_combobox = ttk.Combobox(marco, textvariable=tipoDocumento, values=[texto for id, texto in opciones_tipo_documento])
    tipoDocumento_combobox.bind('<KeyRelease>', lambda event: actualizar_combobox(event, opciones_tipo_documento, tipoDocumento_combobox))

    departamentoExpedicion_combobox = ttk.Combobox(marco, textvariable=departamentoExpedicion, values=[texto for id, texto in opciones_departamento])
    departamentoExpedicion_combobox.bind('<KeyRelease>', lambda event: actualizar_combobox(event, opciones_departamento, departamentoExpedicion_combobox))

    ciudadExpedicion_combobox = ttk.Combobox(marco, textvariable=ciudadExpedicion, values=[texto for id, texto in opciones_ciudad])
    ciudadExpedicion_combobox.bind('<KeyRelease>', lambda event: actualizar_combobox(event, opciones_ciudad, ciudadExpedicion_combobox))

    genero_combobox = ttk.Combobox(marco, textvariable=genero, values=[texto for id, texto in opciones_genero])
    genero_combobox.bind('<KeyRelease>', lambda event: actualizar_combobox(event, opciones_genero, genero_combobox))

    Rh_combobox = ttk.Combobox(marco, textvariable=Rh, values=[texto for id, texto in opciones_rh])
    Rh_combobox.bind('<KeyRelease>', lambda event: actualizar_combobox(event, opciones_rh, Rh_combobox))

    ARL_combobox = ttk.Combobox(marco, textvariable=ARL, values=[texto for id, texto in opciones_arl])
    ARL_combobox.bind('<KeyRelease>', lambda event: actualizar_combobox(event, opciones_arl, ARL_combobox))

    EPS_combobox = ttk.Combobox(marco, textvariable=EPS, values=[texto for id, texto in opciones_eps])
    EPS_combobox.bind('<KeyRelease>', lambda event: actualizar_combobox(event, opciones_eps, EPS_combobox))

    idBanco_combobox = ttk.Combobox(marco, textvariable=idBanco, values=[texto for id, texto in opciones_banco])
    idBanco_combobox.bind('<KeyRelease>', lambda event: actualizar_combobox(event, opciones_banco, idBanco_combobox))

    Estado_combobox = ttk.Combobox(marco, textvariable=Estado, values=[texto for id, texto in opciones_estado])
    Estado_combobox.bind('<KeyRelease>', lambda event: actualizar_combobox(event, opciones_estado, Estado_combobox))

    usuarioCreador_combobox = ttk.Combobox(marco, textvariable=usuarioCreador, values=[texto for id, texto in opciones_usuario_creador])
    usuarioCreador_combobox.bind('<KeyRelease>', lambda event: actualizar_combobox(event, opciones_usuario_creador, usuarioCreador_combobox))

    filas = [
        ("Número Documento", numero_documento, numero_documento_entry),
        ("Primer Nombre", primerNombre, tk.Entry(marco, textvariable=primerNombre, width=20)),
        ("Segundo Nombre", segundoNombre, tk.Entry(marco, textvariable=segundoNombre, width=20)),
        ("Primer Apellido", primerApellido, tk.Entry(marco, textvariable=primerApellido, width=20)),
        ("Segundo Apellido", segundoApellido, tk.Entry(marco, textvariable=segundoApellido, width=20)),
        ("Tipo Documento", tipoDocumento, tipoDocumento_combobox),
        ("Departamento Expedición", departamentoExpedicion, departamentoExpedicion_combobox),
        ("Ciudad Expedición", ciudadExpedicion, ciudadExpedicion_combobox),
        ("Fecha Expedición", fechaExpedicion, DateEntry(marco, textvariable=fechaExpedicion, date_pattern='y-mm-dd')),
        ("Género", genero, genero_combobox),
        ("Fecha Nacimiento", fechaNacimento, DateEntry(marco, textvariable=fechaNacimento, date_pattern='y-mm-dd')),
        ("RH", Rh, Rh_combobox),
        ("Dirección", dirrecion, tk.Entry(marco, textvariable=dirrecion, width=30)),
        ("Correo", correo, tk.Entry(marco, textvariable=correo, width=30)),
        ("Correo Adicional", correoAdicional, tk.Entry(marco, textvariable=correoAdicional, width=30)),
        ("Celular", celular, tk.Entry(marco, textvariable=celular, width=20)),
        ("Teléfono", telefono, tk.Entry(marco, textvariable=telefono, width=20)),
        ("ARL", ARL, ARL_combobox),
        ("EPS", EPS, EPS_combobox),
        ("PAA", PAA, tk.Entry(marco, textvariable=PAA, width=30)),
        ("Banco", idBanco, idBanco_combobox),
        ("Número Cuenta", numeroCuenta, tk.Entry(marco, textvariable=numeroCuenta, width=20)),
        ("CDP", CDP, tk.Entry(marco, textvariable=CDP, width=20)),
        ("Fecha Registro", fechaRegistro, DateEntry(marco, textvariable=fechaRegistro, date_pattern='y-mm-dd')),
        ("Último Acceso", ultimoAcceso, tk.Entry(marco, textvariable=ultimoAcceso, width=20)),
        ("Estado", Estado, Estado_combobox),
        ("Usuario Creador", usuarioCreador, usuarioCreador_combobox)
    ]

    for i, (texto, var, widget) in enumerate(filas):
        columna = i % 3
        fila = i // 3
        tk.Label(marco, text=texto).grid(column=columna*2, row=fila, padx=2, pady=2, sticky="e")
        widget.grid(column=columna*2+1, row=fila, padx=2, pady=2, sticky="ew")

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green", font=("Arial", 12, "bold"))
    lblMensaje.grid(column=0, row=10, columnspan=6, padx=2, pady=2, sticky="ew")

    # Ajustar el texto del mensaje para adaptarse al tamaño de la ventana
    def ajustar_texto_mensaje():
        lblMensaje.config(wraplength=marco.winfo_width())

    marco.bind("<Configure>", lambda event: ajustar_texto_mensaje())

    # Scrollbars
    scrollbar_y = ttk.Scrollbar(marco, orient="vertical")
    scrollbar_y.grid(row=11, column=6, sticky="ns")
    scrollbar_x = ttk.Scrollbar(marco, orient="horizontal")
    scrollbar_x.grid(row=12, column=0, columnspan=6, sticky="ew")

    # Treeview con scrollbars
    tvClientes = ttk.Treeview(marco, yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set, selectmode='browse')
    tvClientes["columns"] = (
        "id_cliente", "primerNombre", "segundoNombre", "primerApellido", "segundoApellido", "tipoDocumento",
        "departamentoExpedicion", "ciudadExpedicion", "fechaExpedicion", "genero", "fechaNacimento", "Rh", "dirrecion", "correo", "correoAdicional", "celular", "telefono",
        "ARL", "EPS", "PAA", "idBanco", "numeroCuenta", "CDP", "fechaRegistro", "ultimoAcceso", "Estado", "usuarioCreador"
    )
    tvClientes.column("#0", width=0, stretch='no')
    tvClientes.column("id_cliente", width=50, anchor='center')

    for col in tvClientes["columns"][1:]:
        tvClientes.column(col, width=80, anchor='center', stretch='yes')
        tvClientes.heading(col, text=col, anchor='center')

    tvClientes.grid(column=0, row=11, columnspan=6, padx=2, pady=2, sticky="nsew")
    scrollbar_y.config(command=tvClientes.yview)
    scrollbar_x.config(command=tvClientes.xview)
    tvClientes.bind("<<TreeviewSelect>>", seleccionar)

    query = ""

    def cargar_datos():
        nonlocal total_paginas, query
        # Consulta SQL con JOIN para obtener los textos correspondientes a las llaves foráneas
        query = """
        SELECT c.id, c.primerNombre, c.segundoNombre, c.primerApellido, c.segundoApellido,
            td.numeroCedula AS tipoDocumento, d.nombreDepartamento AS departamentoExpedicion,
            ci.nombreCiudad AS ciudadExpedicion, c.fechaExpedicion, g.nombreGenero AS genero,
            c.fechaNacimento, r.tipoRh AS Rh, c.dirrecion, c.correo, c.correoAdicional, c.celular, c.telefono,
            a.nombreARL AS ARL, e.nombreEPS AS EPS, c.PAA, b.nombreBanco AS idBanco,
            c.numeroCuenta, c.CDP, c.fechaRegistro, c.ultimoAcceso, es.tipoEstado AS Estado, u.numeroCedulaCliente AS usuarioCreador
        FROM clientes c
        JOIN tipodedocumento td ON c.tipoDocumento = td.id
        JOIN departamentos d ON c.departamentoExpedicion = d.id
        JOIN ciudad ci ON c.ciudadExpedicion = ci.id
        JOIN genero g ON c.genero = g.id
        JOIN rh r ON c.Rh = r.id
        JOIN arl a ON c.ARL = a.id
        JOIN eps e ON c.EPS = e.id
        JOIN banco b ON c.idBanco = b.id
        JOIN estado es ON c.Estado = es.id
        JOIN usuarios u ON c.usuarioCreador = u.id
        """
        cursor.execute(query)
        registros = cursor.fetchall()
        total_registros = len(registros)
        total_paginas = (total_registros + registros_por_pagina - 1) // registros_por_pagina
        mostrar_pagina(pagina_actual)

    def mostrar_pagina(pagina):
        tvClientes.delete(*tvClientes.get_children())
        offset = pagina * registros_por_pagina
        paginated_query = query + f" LIMIT {registros_por_pagina} OFFSET {offset}"
        cursor.execute(paginated_query)
        registros = cursor.fetchall()

        for registro in registros:
            tvClientes.insert("", "end", values=registro)
    def avanzar_pagina():
        nonlocal pagina_actual
        if pagina_actual < total_paginas - 1:
            pagina_actual += 1
            mostrar_pagina(pagina_actual)

    def retroceder_pagina():
        nonlocal pagina_actual
        if pagina_actual > 0:
            pagina_actual -= 1
            mostrar_pagina(pagina_actual)

    def obtener_id_por_texto(opciones, texto):
        for id, txt in opciones:
            if str(txt) == str(texto):  # Asegurarse de comparar los valores como cadenas
                return id
        return None

    def exportar_excel():
        archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if archivo:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Clientes"

            for i, col in enumerate(tvClientes["columns"], 1):
                ws.cell(row=1, column=i, value=col)

            for i, item in enumerate(tvClientes.get_children(), 2):
                for j, valor in enumerate(tvClientes.item(item)['values'], 1):
                    ws.cell(row=i, column=j, value=valor)

            wb.save(archivo)
            lblMensaje.config(text=f"Datos exportados a {archivo}", fg="green", font=("Arial", 12, "bold"))
            lblMensaje.grid(columnspan=6, sticky="ew")

    def exportar_pdf():
        archivo = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if archivo:
            c = canvas.Canvas(archivo, pagesize=letter)
            width, height = letter

            y = height - 40
            for i, col in enumerate(tvClientes["columns"], 1):
                c.drawString(10 * i, y, col)

            y -= 20
            for item in tvClientes.get_children():
                for j, valor in enumerate(tvClientes.item(item)['values'], 1):
                    c.drawString(10 * j, y, str(valor))
                y -= 20
                if y < 40:
                    c.showPage()
                    y = height - 40

            c.save()
            lblMensaje.config(text=f"Datos exportados a {archivo}", fg="green", font=("Arial", 12, "bold"))
            lblMensaje.grid(columnspan=6, sticky="ew")
            
    def agregar_cliente():
        if not validar_campos():
            return

        # Obtener los IDs correspondientes a los textos seleccionados en los comboboxes
        id_tipo_documento = obtener_id_por_texto(opciones_tipo_documento, tipoDocumento.get())
        id_departamento = obtener_id_por_texto(opciones_departamento, departamentoExpedicion.get())
        id_ciudad = obtener_id_por_texto(opciones_ciudad, ciudadExpedicion.get())
        id_genero = obtener_id_por_texto(opciones_genero, genero.get())
        id_rh = obtener_id_por_texto(opciones_rh, Rh.get())
        id_arl = obtener_id_por_texto(opciones_arl, ARL.get())
        id_eps = obtener_id_por_texto(opciones_eps, EPS.get())
        id_banco = obtener_id_por_texto(opciones_banco, idBanco.get())
        id_estado = obtener_id_por_texto(opciones_estado, Estado.get())
        id_usuario_creador = obtener_id_por_texto(opciones_usuario_creador, usuarioCreador.get())

        if id_usuario_creador is None:
            lblMensaje.config(text="Error: No se ha seleccionado un usuario válido en 'Usuario Creador'.", fg="red", font=("Arial", 12, "bold"))
            lblMensaje.grid(columnspan=6, sticky="ew")
            return

        # Generar la tupla de valores con exactamente 27 elementos
        valores = (
            numero_documento.get(), primerNombre.get(), segundoNombre.get(), primerApellido.get(), segundoApellido.get(),
            id_tipo_documento, id_departamento, id_ciudad, fechaExpedicion.get(), id_genero,
            fechaNacimento.get(), id_rh, dirrecion.get(), correo.get(), correoAdicional.get(), celular.get(), telefono.get(), id_arl, id_eps,
            PAA.get(), id_banco, numeroCuenta.get(), CDP.get(), fechaRegistro.get(), ultimoAcceso.get(), id_estado, id_usuario_creador,  # Aquí añadimos el valor extra
        )

        # Confirmar que la tupla tiene exactamente 27 valores
        if len(valores) != 27:
            print(f"Error: Número de valores no coincide con los placeholders. Valores: {len(valores)}")
            return

        # Consulta SQL
        query = """
        INSERT INTO clientes (id, primerNombre, segundoNombre, primerApellido, segundoApellido, tipoDocumento,
                            departamentoExpedicion, ciudadExpedicion, fechaExpedicion, genero, fechaNacimento, Rh, dirrecion,
                            correo, correoAdicional, celular, telefono, ARL, EPS, PAA, idBanco, numeroCuenta, CDP, fechaRegistro, ultimoAcceso, Estado, usuarioCreador)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        try:
            # Ejecutar la consulta
            cursor.execute(query, valores)
            db.commit()
            lblMensaje.config(text="Cliente agregado exitosamente", fg="green", font=("Arial", 12, "bold"))
            lblMensaje.grid(columnspan=6, sticky="ew")
            cargar_datos()
        except mysql.connector.Error as err:
            # Imprimir el error en la consola
            print("Error durante la ejecución de la consulta SQL:", err)
            print("Consulta SQL:", query)
            print("Valores:", valores)

    def actualizar_cliente():
        if not validar_campos():
            return

        id_tipo_documento = obtener_id_por_texto(opciones_tipo_documento, tipoDocumento.get())
        id_departamento = obtener_id_por_texto(opciones_departamento, departamentoExpedicion.get())
        id_ciudad = obtener_id_por_texto(opciones_ciudad, ciudadExpedicion.get())
        id_genero = obtener_id_por_texto(opciones_genero, genero.get())
        id_rh = obtener_id_por_texto(opciones_rh, Rh.get())
        id_arl = obtener_id_por_texto(opciones_arl, ARL.get())
        id_eps = obtener_id_por_texto(opciones_eps, EPS.get())
        id_banco = obtener_id_por_texto(opciones_banco, idBanco.get())
        id_estado = obtener_id_por_texto(opciones_estado, Estado.get())
        id_usuario_creador = obtener_id_por_texto(opciones_usuario_creador, usuarioCreador.get())

        if id_usuario_creador is None:
            lblMensaje.config(text="Error: No se ha seleccionado un usuario válido en 'Usuario Creador'.", fg="red", font=("Arial", 12, "bold"))
            lblMensaje.grid(columnspan=6, sticky="ew")
            return

        query = """
        UPDATE clientes SET primerNombre=%s, segundoNombre=%s, primerApellido=%s, segundoApellido=%s, tipoDocumento=%s,
                            departamentoExpedicion=%s, ciudadExpedicion=%s, fechaExpedicion=%s, genero=%s, fechaNacimento=%s,
                            Rh=%s, dirrecion=%s, correo=%s, correoAdicional=%s, celular=%s, telefono=%s, ARL=%s, EPS=%s, PAA=%s, idBanco=%s,
                            numeroCuenta=%s, CDP=%s, fechaRegistro=%s, ultimoAcceso=%s, Estado=%s, usuarioCreador=%s
        WHERE id=%s
        """
        valores = (
            primerNombre.get(), segundoNombre.get(), primerApellido.get(), segundoApellido.get(), id_tipo_documento,
            id_departamento, id_ciudad, fechaExpedicion.get(), id_genero, fechaNacimento.get(),
            id_rh, dirrecion.get(), correo.get(), correoAdicional.get(), celular.get(), telefono.get(), id_arl, id_eps, PAA.get(),
            id_banco, numeroCuenta.get(), CDP.get(), fechaRegistro.get(), ultimoAcceso.get(), id_estado, id_usuario_creador,
            id_cliente.get()
        )
        cursor.execute(query, valores)
        db.commit()
        lblMensaje.config(text="Cliente actualizado exitosamente", fg="green", font=("Arial", 12, "bold"))
        lblMensaje.grid(columnspan=6, sticky="ew")
        cargar_datos()

    def eliminar_cliente():
        query = "DELETE FROM clientes WHERE id=%s"
        cursor.execute(query, (id_cliente.get(),))
        db.commit()
        lblMensaje.config(text="Cliente eliminado exitosamente", fg="green", font=("Arial", 12, "bold"))
        lblMensaje.grid(columnspan=6, sticky="ew")
        cargar_datos()

    # Botones de acción
    btnAgregar = tk.Button(marco, text="Agregar", command=agregar_cliente)
    btnAgregar.grid(column=0, row=13, padx=5, pady=5)
    btnActualizar = tk.Button(marco, text="Actualizar", command=actualizar_cliente)
    btnActualizar.grid(column=1, row=13, padx=5, pady=5)
    btnEliminar = tk.Button(marco, text="Eliminar", command=eliminar_cliente)
    btnEliminar.grid(column=2, row=13, padx=5, pady=5)
    btnRetroceder = tk.Button(marco, text="<<", command=retroceder_pagina)
    btnRetroceder.grid(column=3, row=13, padx=5, pady=5)
    btnAvanzar = tk.Button(marco, text=">>", command=avanzar_pagina)
    btnAvanzar.grid(column=4, row=13, padx=5, pady=5)
    btnLimpiar = tk.Button(marco, text="Limpiar", command=limpiar_campos)
    btnLimpiar.grid(column=5, row=13, padx=5, pady=5)
    btnExportarExcel = tk.Button(marco, text="Exportar Excel", command=exportar_excel)
    btnExportarExcel.grid(column=6, row=13, padx=5, pady=5)
    btnExportarPDF = tk.Button(marco, text="Exportar PDF", command=exportar_pdf)
    btnExportarPDF.grid(column=7, row=13, padx=5, pady=5)

    cargar_datos()

def mostrar_jefes(ventana):
    global db
    db = DataBase()
    global page_number, records_per_page, total_records
    modificar = False
    id_jefe = tk.StringVar()
    nombre_jefe = tk.StringVar()

    # Inicializar variables de paginación
    page_number = 1
    records_per_page = 10
    total_records = 0

    def seleccionar(event):
        seleccion = tvJefes.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvJefes.item(id, "values")
            id_jefe.set(valores[0])
            nombre_jefe.set(valores[1])

    marco = tk.LabelFrame(ventana, text="Formulario Jefes")
    marco.place(x=50, y=50, width=500, height=400)

    tk.Label(marco, text="ID").grid(column=0, row=0, padx=5, pady=5)
    txtIdJefe = tk.Entry(marco, textvariable=id_jefe)
    txtIdJefe.grid(column=1, row=0)

    tk.Label(marco, text="Nombre").grid(column=0, row=1, padx=5, pady=5)
    txtNombreJefe = tk.Entry(marco, textvariable=nombre_jefe)
    txtNombreJefe.grid(column=1, row=1)

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=2, columnspan=4)

    tvJefes = ttk.Treeview(marco, selectmode='none')
    tvJefes["columns"] = ("ID", "Nombre")
    tvJefes.column("#0", width=0, stretch='no')
    tvJefes.column("ID", width=150, anchor='center')
    tvJefes.column("Nombre", width=150, anchor='center')
    tvJefes.heading("#0", text="")
    tvJefes.heading("ID", text="ID", anchor='center')
    tvJefes.heading("Nombre", text="Nombre", anchor='center')
    tvJefes.grid(column=0, row=3, columnspan=4, padx=5)
    tvJefes.bind("<<TreeviewSelect>>", seleccionar)

    btnEliminar = tk.Button(marco, text="Eliminar", command=lambda: eliminar())
    btnEliminar.grid(column=1, row=4)
    btnNuevo = tk.Button(marco, text="Guardar", command=lambda: nuevo())
    btnNuevo.grid(column=2, row=4)
    btnModificar = tk.Button(marco, text="Seleccionar", command=lambda: actualizar())
    btnModificar.grid(column=3, row=4)

    btnCerrar = tk.Button(marco, text="X", command=ventana.destroy, bg="red", fg="white")
    btnCerrar.place(x=470, y=10, width=20, height=20)
    
    btnMinimizar = tk.Button(marco, text="-", command=lambda: ventana.iconify(), bg="yellow", fg="black")
    btnMinimizar.place(x=450, y=10, width=20, height=20)

    # Botones de paginación
    def paginacion():
        llenar_tabla()
        # Actualizar el estado de los botones de navegación
        btnPrev.config(state='normal' if page_number > 1 else 'disabled')
        btnNext.config(state='normal' if page_number * records_per_page < total_records else 'disabled')

    def prev_page():
        global page_number
        if page_number > 1:
            page_number -= 1
            llenar_tabla()

    def next_page():
        global page_number
        if page_number * records_per_page < total_records:
            page_number += 1
            llenar_tabla()

    btnPrev = tk.Button(marco, text="<< Anterior", command=prev_page)
    btnPrev.grid(column=0, row=5, pady=10, sticky='w')
    
    btnNext = tk.Button(marco, text="Siguiente >>", command=next_page)
    btnNext.grid(column=3, row=5, pady=10, sticky='e')

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tvJefes.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Seleccionar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tvJefes.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def validar():
        return len(id_jefe.get()) > 0 and len(nombre_jefe.get()) > 0

    def limpiar():
        id_jefe.set("")
        nombre_jefe.set("")

    def vaciar_tabla():
        filas = tvJefes.get_children()
        for fila in filas:
            tvJefes.delete(fila)

    def llenar_tabla():
        vaciar_tabla()
        global total_records
        offset = (page_number - 1) * records_per_page
        sql_count = "SELECT COUNT(*) FROM jefe"
        db.cursor.execute(sql_count)
        total_records = db.cursor.fetchone()[0]
        
        sql = "SELECT id, nombreJefe FROM jefe LIMIT %s OFFSET %s"
        db.cursor.execute(sql, (records_per_page, offset))
        filas = db.cursor.fetchall()
        for fila in filas:
            id = fila[0]
            tvJefes.insert("", 'end', id, text=id, values=(fila[0], fila[1]))
        # Actualizar el estado de los botones de navegación
        btnPrev.config(state='normal' if page_number > 1 else 'disabled')
        btnNext.config(state='normal' if page_number * records_per_page < total_records else 'disabled')

    def eliminar():
        seleccion = tvJefes.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM jefe WHERE id=%s"
                db.cursor.execute(sql, (id,))
                db.connection.commit()
                tvJefes.delete(id)
                lblMensaje.config(text="Se ha eliminado el registro correctamente")
                limpiar()
                llenar_tabla()
            else:
                lblMensaje.config(text="Seleccione un registro para eliminar")

    def nuevo():
        if not modificar:
            if validar():
                val = (id_jefe.get(), nombre_jefe.get())
                sql = "INSERT INTO jefe (id, nombreJefe) VALUES (%s, %s)"
                db.cursor.execute(sql, val)
                db.connection.commit()
                lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                llenar_tabla()
                limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarFalse()

    def actualizar():
        if modificar:
            if validar():
                seleccion = tvJefes.selection()
                if seleccion:
                    id = seleccion[0]
                    val = (id_jefe.get(), nombre_jefe.get())
                    sql = "UPDATE jefe SET id=%s, nombreJefe=%s WHERE id=%s"
                    db.cursor.execute(sql, val + (id,))
                    db.connection.commit()
                    lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                    llenar_tabla()
                    limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarTrue()

    llenar_tabla()

def mostrar_eps(ventana):
    global db
    db = DataBase()
    modificar = False
    dni = tk.StringVar()
    nombre = tk.StringVar()

    def exportar_a_excel_eps():
        try:
            sql = "SELECT id, nombreEPS FROM eps"
            db.cursor.execute(sql)
            filas = db.cursor.fetchall()
            if filas:
                df = pd.DataFrame(filas, columns=['ID', 'Nombre'])

                filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                        filetypes=[("Excel files", "*.xlsx"), 
                                                                   ("All files", "*.*")])
                
                if filepath:
                    df.to_excel(filepath, index=False)
                    messagebox.showinfo("Exportación Exitosa", f"Datos exportados a {filepath}")
                    os.startfile(filepath)
                else:
                    messagebox.showwarning("Cancelado", "Exportación cancelada por el usuario")
            else:
                messagebox.showwarning("Sin Datos", "No hay datos para exportar")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar los datos: {e}")

    def importar_de_excel_eps():
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if filepath:
            try:
                df = pd.read_excel(filepath)
                for i, row in df.iterrows():
                    sql = "INSERT INTO eps (id, nombreEPS) VALUES (%s, %s)"
                    db.cursor.execute(sql, tuple(row))
                db.connection.commit()
                messagebox.showinfo("Importación Exitosa", "Datos importados correctamente desde el archivo Excel")
                llenar_tabla()  # Recargar la tabla después de importar
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo importar los datos: {e}")
        else:
            messagebox.showwarning("Cancelado", "Importación cancelada por el usuario")

    def seleccionar(event):
        seleccion = tveps.selection()
        if seleccion:
            id = seleccion[0]
            valores = tveps.item(id, "values")
            dni.set(valores[0])
            nombre.set(valores[1])

    marco = tk.LabelFrame(ventana, text="Formulario EPS")
    marco.place(x=50, y=50, width=500, height=400)

    tk.Label(marco, text="DNI").grid(column=0, row=0, padx=5, pady=5)
    txtDni = tk.Entry(marco, textvariable=dni)
    txtDni.grid(column=1, row=0)

    tk.Label(marco, text="Nombre").grid(column=0, row=1, padx=5, pady=5)
    txtNombre = tk.Entry(marco, textvariable=nombre)
    txtNombre.grid(column=1, row=1)

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=2, columnspan=4)

    tveps = ttk.Treeview(marco, selectmode='none')
    tveps["columns"] = ("DNI", "Nombre")
    tveps.column("#0", width=0, stretch='no')
    tveps.column("DNI", width=150, anchor='center')
    tveps.column("Nombre", width=150, anchor='center')
    tveps.heading("#0", text="")
    tveps.heading("DNI", text="DNI", anchor='center')
    tveps.heading("Nombre", text="Nombre", anchor='center')
    tveps.grid(column=0, row=3, columnspan=4, padx=5)
    tveps.bind("<<TreeviewSelect>>", seleccionar)

    btnEliminar = tk.Button(marco, text="Eliminar", command=lambda: eliminar())
    btnEliminar.grid(column=1, row=4)
    btnNuevo = tk.Button(marco, text="Guardar", command=lambda: nuevo())
    btnNuevo.grid(column=2, row=4)
    btnModificar = tk.Button(marco, text="Seleccionar", command=lambda: actualizar())
    btnModificar.grid(column=3, row=4)

    # Botones para exportar e importar datos a Excel
    btnExportar = tk.Button(marco, text="Exportar a Excel", command=exportar_a_excel_eps)
    btnExportar.grid(column=2, row=5, pady=10)
    btnImportar = tk.Button(marco, text="Importar de Excel", command=importar_de_excel_eps)
    btnImportar.grid(column=3, row=5, pady=10)

    # Botón para cerrar y minimizar
    btnCerrar = tk.Button(marco, text="X", command=ventana.destroy, bg="red", fg="white")
    btnCerrar.place(x=470, y=10, width=20, height=20)
    
    btnMinimizar = tk.Button(marco, text="-", command=lambda: ventana.iconify(), bg="yellow", fg="black")
    btnMinimizar.place(x=450, y=10, width=20, height=20)

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tveps.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Seleccionar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tveps.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def validar():
        return len(dni.get()) > 0 and len(nombre.get()) > 0

    def limpiar():
        dni.set("")
        nombre.set("")

    def vaciar_tabla():
        filas = tveps.get_children()
        for fila in filas:
            tveps.delete(fila)

    def llenar_tabla():
        vaciar_tabla()
        sql = "SELECT id, nombreEPS FROM eps"
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        for fila in filas:
            id = fila[0]
            tveps.insert("", 'end', id, text=id, values=(fila[0], fila[1]))

    def eliminar():
        seleccion = tveps.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM eps WHERE id=%s"
                db.cursor.execute(sql, (id,))
                db.connection.commit()
                tveps.delete(id)
                lblMensaje.config(text="Se ha eliminado el registro correctamente")
                limpiar()
            else:
                lblMensaje.config(text="Seleccione un registro para eliminar")

    def nuevo():
        if not modificar:
            if validar():
                val = (dni.get(), nombre.get())
                sql = "INSERT INTO eps (id, nombreEPS) VALUES (%s, %s)"
                db.cursor.execute(sql, val)
                db.connection.commit()
                lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                llenar_tabla()
                limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarFalse()

    def actualizar():
        if modificar:
            if validar():
                seleccion = tveps.selection()
                if seleccion:
                    id = seleccion[0]
                    val = (dni.get(), nombre.get())
                    sql = "UPDATE eps SET id=%s, nombreEPS=%s WHERE id=%s"
                    db.cursor.execute(sql, val + (id,))
                    db.connection.commit()
                    lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                    llenar_tabla()
                    limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarTrue()

    llenar_tabla()

def exportar_a_excel():
    try:
        sql = """
        SELECT b.id, b.nombreBanco, tc.nombreTipoCuenta
        FROM banco b
        JOIN tipodecuenta tc ON b.tipoCuenta = tc.id
        """
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        if filas:
            df = pd.DataFrame(filas, columns=['ID Banco', 'Nombre Banco', 'Tipo Cuenta'])

            filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
            
            if filepath:
                df.to_excel(filepath, index=False)
                messagebox.showinfo("Exportación Exitosa", f"Datos exportados a {filepath}")
                os.startfile(filepath)
            else:
                messagebox.showwarning("Cancelado", "Exportación cancelada por el usuario")
        else:
            messagebox.showwarning("Sin Datos", "No hay datos para exportar")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar los datos: {e}")

def importar_de_excel():
    try:
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        
        if filepath:
            df = pd.read_excel(filepath)
            
            if 'ID Banco' in df.columns and 'Nombre Banco' in df.columns and 'Tipo Cuenta' in df.columns:
                for _, row in df.iterrows():
                    id_banco = row['ID Banco']
                    nombre_banco = row['Nombre Banco']
                    tipo_cuenta = row['Tipo Cuenta']

                    if tipo_cuenta == 'Debito':
                        tipo_cuenta = 1
                    elif tipo_cuenta == 'Credito':
                        tipo_cuenta = 2
                    else:
                        # Si hay un tipo de cuenta desconocido, puedes omitir el registro o manejar el error
                        print(f"Tipo de cuenta desconocido: {tipo_cuenta}")
                        continue
                    
                    try:
                        # Verificar valores antes de la inserción
                        print(f"Inserción: ID={id_banco}, Nombre={nombre_banco}, Tipo Cuenta={tipo_cuenta}")
                        sql = "INSERT INTO banco (id, nombreBanco, tipoCuenta) VALUES (%s, %s, %s)"
                        db.cursor.execute(sql, (id_banco, nombre_banco, tipo_cuenta))
                    except Exception as e:
                        print(f"Error al insertar el registro con ID {id_banco}: {e}")
                        continue
                
                db.connection.commit()
                messagebox.showinfo("Importación Exitosa", "Datos importados correctamente desde el archivo Excel")
                llenar_tabla()  # Refresca la tabla con los nuevos datos importados
            else:
                messagebox.showerror("Error", "El archivo Excel no tiene las columnas correctas ('ID Banco', 'Nombre Banco', 'Tipo Cuenta')")
        else:
            messagebox.showwarning("Cancelado", "Importación cancelada por el usuario")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo importar los datos: {e}")
def mostrar_bancos(ventana):
    global db
    global llenar_tabla
    db = DataBase()
    global modificar
    modificar = False

    # Variables de paginación
    page_number = 1
    records_per_page = 10

    # Variables de filtro
    banco_seleccionado = tk.StringVar()
    tipo_cuenta_seleccionada = tk.StringVar()

    def seleccionar(event):
        seleccion = tvBancos.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvBancos.item(id, "values")
            dni.set(valores[0])
            nombre.set(valores[1])
            tipoCuenta.set(valores[2])
            # Disable ID field during modification
            txtID.config(state='disabled')

    def cargar_tipo_cuenta():
        sql = "SELECT nombreTipoCuenta FROM tipodecuenta"
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        return [fila[0] for fila in filas]

    def cargar_bancos():
        sql = "SELECT DISTINCT nombreBanco FROM banco"
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        return [fila[0] for fila in filas]

    def vaciar_tabla():
        filas = tvBancos.get_children()
        for fila in filas:
            tvBancos.delete(fila)

    def llenar_tabla(filtro_banco=None, filtro_tipo_cuenta=None):
        vaciar_tabla()
        try:
            offset = (page_number - 1) * records_per_page
            sql = """
            SELECT b.id, b.nombreBanco, tc.nombreTipoCuenta
            FROM banco b
            JOIN tipodecuenta tc ON b.tipoCuenta = tc.id
            """
            params = []

            # Filtrar por banco y tipo de cuenta
            if filtro_banco:
                sql += " WHERE b.nombreBanco = %s"
                params.append(filtro_banco)
            if filtro_tipo_cuenta:
                sql += " AND tc.nombreTipoCuenta = %s" if filtro_banco else " WHERE tc.nombreTipoCuenta = %s"
                params.append(filtro_tipo_cuenta)

            sql += " LIMIT %s OFFSET %s"
            params.extend([records_per_page, offset])
            
            db.cursor.execute(sql, tuple(params))
            filas = db.cursor.fetchall()
            for fila in filas:
                id = fila[0]
                tvBancos.insert("", 'end', id, text=id, values=(fila[0], fila[1], fila[2]))
        except Exception as e:
            lblMensaje.config(text=f"Error al llenar la tabla: {e}", fg="red")

    def filtrar_registros(event=None):
        filtro_banco = banco_seleccionado.get()
        filtro_tipo_cuenta = tipo_cuenta_seleccionada.get()
        llenar_tabla(filtro_banco, filtro_tipo_cuenta)

    def mostrar_todos():
        banco_seleccionado.set("")
        tipo_cuenta_seleccionada.set("")
        llenar_tabla()

    def limpiar():
        dni.set("")
        nombre.set("")
        tipoCuenta.set("")
        # Enable ID field when not modifying
        txtID.config(state='normal')

    
    def eliminar():
        seleccion = tvBancos.selection()
        if seleccion:
            id = seleccion[0]
            sql = "DELETE FROM banco WHERE id=%s"
            db.cursor.execute(sql, (id,))
            db.connection.commit()
            lblMensaje.config(text="Se ha eliminado el registro correctamente", fg="green")
            llenar_tabla()
            limpiar()
        else:
            lblMensaje.config(text="Seleccione un registro para eliminar", fg="red")

    def nuevo():
        if validar():
            tipo_cuenta_id = obtener_tipo_cuenta_id(tipoCuenta.get())
            if tipo_cuenta_id is not None:
                val = (dni.get(), nombre.get(), tipo_cuenta_id)
                sql = "INSERT INTO banco (id, nombreBanco, tipoCuenta) VALUES (%s, %s, %s)"
                db.cursor.execute(sql, val)
                db.connection.commit()
                lblMensaje.config(text="Se ha guardado el registro correctamente", fg="green")
                llenar_tabla()
                limpiar()
            else:
                lblMensaje.config(text="Tipo de cuenta inválido", fg="red")
        else:
            lblMensaje.config(text="Todos los campos son requeridos", fg="red")
    def validar():
        # Aquí comprueba que los campos necesarios estén completos y que el tipo de cuenta sea válido.
        return len(dni.get()) > 0 and len(nombre.get()) > 0 and tipoCuenta.get() in cargar_tipo_cuenta()

    def actualizar():
        global modificar  # Usamos global para modificar la variable global
        if modificar:
            seleccion = tvBancos.selection()
            if seleccion:
                id = seleccion[0]
                if validar():  # Aquí se llama la función validar()
                    tipo_cuenta_id = obtener_tipo_cuenta_id(tipoCuenta.get())
                    if tipo_cuenta_id is not None:
                        val = (nombre.get(), tipo_cuenta_id, dni.get())
                        sql = "UPDATE banco SET nombreBanco=%s, tipoCuenta=%s WHERE id=%s"
                        db.cursor.execute(sql, val)
                        db.connection.commit()
                        lblMensaje.config(text="Se ha actualizado el registro correctamente", fg="green")
                        llenar_tabla()
                        limpiar()
                        modificar = False  # Cambia el valor de modificar a False
                    else:
                        lblMensaje.config(text="Tipo de cuenta inválido", fg="red")
                else:
                    lblMensaje.config(text="Todos los campos son requeridos", fg="red")
            else:
                lblMensaje.config(text="Seleccione un registro para modificar", fg="red")
        else:
            modificar = True 

    def obtener_tipo_cuenta_id(nombre_tipo_cuenta):
        sql = "SELECT id FROM tipodecuenta WHERE nombreTipoCuenta=%s"
        db.cursor.execute(sql, (nombre_tipo_cuenta,))
        resultado = db.cursor.fetchone()
        return resultado[0] if resultado else None

    def ir_a_pagina_anterior():
        nonlocal page_number
        if page_number > 1:
            page_number -= 1
            llenar_tabla()
            actualizar_botones_paginacion()

    def ir_a_pagina_siguiente():
        nonlocal page_number
        page_number += 1
        llenar_tabla()
        actualizar_botones_paginacion()

    def actualizar_botones_paginacion():
        btnAnterior.config(state='normal' if page_number > 1 else 'disabled')
        btnSiguiente.config(state='normal')  # Logic for disabling based on records can be added here

    # Configuración del marco y los widgets
    marco = tk.LabelFrame(ventana, text="Formulario Bancos")
    marco.place(x=50, y=50, width=600, height=500)

    dni = tk.StringVar()
    nombre = tk.StringVar()
    tipoCuenta = tk.StringVar()

    # Selectores de Banco y Tipo Cuenta
    tk.Label(marco, text="Banco").grid(column=0, row=0, padx=5, pady=5)
    comboBanco = ttk.Combobox(marco, textvariable=banco_seleccionado)
    comboBanco['values'] = cargar_bancos()
    comboBanco.grid(column=1, row=0)
    comboBanco.bind("<<ComboboxSelected>>", filtrar_registros)  # Apply filter automatically on selection

    tk.Label(marco, text="Tipo Cuenta").grid(column=0, row=1, padx=5, pady=5)
    comboTipoCuentaFiltro = ttk.Combobox(marco, textvariable=tipo_cuenta_seleccionada)
    comboTipoCuentaFiltro['values'] = cargar_tipo_cuenta()
    comboTipoCuentaFiltro.grid(column=1, row=1)
    comboTipoCuentaFiltro.bind("<<ComboboxSelected>>", filtrar_registros)  # Apply filter automatically on selection

    # Botón de Mostrar Todos
    btnMostrarTodos = tk.Button(marco, text="Mostrar Todos", command=mostrar_todos)
    btnMostrarTodos.grid(column=2, row=1, padx=10)

    # Campos del formulario para ID, Nombre, Tipo de Cuenta
    tk.Label(marco, text="ID Banco").grid(column=0, row=2, padx=5, pady=5)
    txtID = tk.Entry(marco, textvariable=dni)
    txtID.grid(column=1, row=2)

    tk.Label(marco, text="Nombre Banco").grid(column=0, row=3, padx=5, pady=5)
    txtNombre = tk.Entry(marco, textvariable=nombre)
    txtNombre.grid(column=1, row=3)

    tk.Label(marco, text="Tipo Cuenta").grid(column=0, row=4, padx=5, pady=5)
    comboTipoCuenta = ttk.Combobox(marco, textvariable=tipoCuenta)
    comboTipoCuenta.grid(column=1, row=4)
    comboTipoCuenta['values'] = cargar_tipo_cuenta()
    comboTipoCuenta.set('')

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=5, columnspan=4)

    # Botones de acción
    btnNuevo = tk.Button(marco, text="Guardar", command=nuevo)
    btnNuevo.grid(column=0, row=6, padx=5, pady=5)

    btnModificar = tk.Button(marco, text="Modificar", command=actualizar)
    btnModificar.grid(column=1, row=6, padx=5, pady=5)

    btnEliminar = tk.Button(marco, text="Eliminar", command=eliminar)
    btnEliminar.grid(column=2, row=6, padx=5, pady=5)

    btnLimpiar = tk.Button(marco, text="Limpiar", command=limpiar)
    btnLimpiar.grid(column=3, row=6, padx=5, pady=5)

    # Paginación
    btnAnterior = tk.Button(marco, text="Anterior", command=ir_a_pagina_anterior, state='disabled')
    btnAnterior.grid(column=0, row=7, padx=5, pady=5)

    btnSiguiente = tk.Button(marco, text="Siguiente", command=ir_a_pagina_siguiente)
    btnSiguiente.grid(column=1, row=7, padx=5, pady=5)

    # Tabla
    tvBancos = ttk.Treeview(marco, columns=("ID", "Nombre", "Tipo Cuenta"), show='headings')
    tvBancos.column("ID", width=50)
    tvBancos.column("Nombre", width=150)
    tvBancos.column("Tipo Cuenta", width=100)

    tvBancos.heading("ID", text="ID")
    tvBancos.heading("Nombre", text="Nombre Banco")
    tvBancos.heading("Tipo Cuenta", text="Tipo de Cuenta")

    tvBancos.grid(column=0, row=8, columnspan=4, padx=5, pady=5)

    # Configurar selección de fila mediante clic de mouse
    tvBancos.bind("<<TreeviewSelect>>", seleccionar)

    # Inicializar la tabla con datos
    llenar_tabla()

    # Actualizar botones de paginación
    actualizar_botones_paginacion()

def mostrar_ciudad(ventana):
    global db, page_number, records_per_page, total_records
    db = DataBase()
    modificar = False
    dni = tk.StringVar()
    nombre = tk.StringVar()
    filtro_id = tk.StringVar()
    filtro_nombre = tk.StringVar()

    # Inicializar variables de paginación
    page_number = 1
    records_per_page = 10
    total_records = 0

    def seleccionar(event):
        seleccion = tvbancos.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvbancos.item(id, "values")
            dni.set(valores[0])
            nombre.set(valores[1])

    def vaciar_tabla():
        filas = tvbancos.get_children()
        for fila in filas:
            tvbancos.delete(fila)

    def llenar_tabla():
        vaciar_tabla()
        global total_records
        offset = (page_number - 1) * records_per_page
        try:
            # Construir la consulta SQL con filtros
            sql_count = "SELECT COUNT(*) FROM ciudad WHERE 1=1"
            sql = "SELECT id, nombreCiudad FROM ciudad WHERE 1=1"
            parametros = []

            if filtro_id.get():
                sql_count += " AND id LIKE %s"
                sql += " AND id LIKE %s"
                parametros.append(f"%{filtro_id.get()}%")

            if filtro_nombre.get():
                sql_count += " AND nombreCiudad LIKE %s"
                sql += " AND nombreCiudad LIKE %s"
                parametros.append(f"%{filtro_nombre.get()}%")

            # Contar total de registros filtrados
            db.cursor.execute(sql_count, tuple(parametros))
            total_records = db.cursor.fetchone()[0]

            # Agregar límites para paginación
            sql += " LIMIT %s OFFSET %s"
            parametros.extend([records_per_page, offset])
            db.cursor.execute(sql, tuple(parametros))
            filas = db.cursor.fetchall()

            for fila in filas:
                tvbancos.insert("", 'end', iid=fila[0], text=fila[0], values=(fila[0], fila[1]))

            # Actualizar el estado de los botones de navegación
            btnPrev.config(state='normal' if page_number > 1 else 'disabled')
            btnNext.config(state='normal' if page_number * records_per_page < total_records else 'disabled')
        except Exception as e:
            lblMensaje.config(text=f"Error al llenar la tabla: {e}", fg="red")

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tvbancos.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Seleccionar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tvbancos.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def validar():
        return len(dni.get()) > 0 and len(nombre.get()) > 0

    def limpiar():
        dni.set("")
        nombre.set("")

    def eliminar():
        seleccion = tvbancos.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM ciudad WHERE id=%s"
                db.cursor.execute(sql, (id,))
                db.connection.commit()
                tvbancos.delete(id)
                lblMensaje.config(text="Se ha eliminado el registro correctamente", fg="green")
                limpiar()
                llenar_tabla()
            else:
                lblMensaje.config(text="Seleccione un registro para eliminar", fg="red")

    def nuevo():
        if not modificar:
            if validar():
                val = (nombre.get(),)
                sql = "INSERT INTO ciudad (nombreCiudad) VALUES (%s)"
                try:
                    db.cursor.execute(sql, val)
                    db.connection.commit()
                    lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                    llenar_tabla()
                    limpiar()
                except Exception as e:
                    lblMensaje.config(text=f"Error: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarFalse()

    def actualizar():
        if modificar:
            if validar():
                seleccion = tvbancos.selection()
                if seleccion:
                    id = seleccion[0]
                    val = (nombre.get(), id)
                    sql = "UPDATE ciudad SET nombreCiudad=%s WHERE id=%s"
                    try:
                        db.cursor.execute(sql, val)
                        db.connection.commit()
                        lblMensaje.config(text="Se ha actualizado el registro con éxito", fg="green")
                        llenar_tabla()
                        limpiar()
                    except Exception as e:
                        lblMensaje.config(text=f"Error: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarTrue()

    # Botones de paginación
    def paginacion():
        llenar_tabla()
        btnPrev.config(state='normal' if page_number > 1 else 'disabled')
        btnNext.config(state='normal' if page_number * records_per_page < total_records else 'disabled')

    def prev_page():
        global page_number
        if page_number > 1:
            page_number -= 1
            llenar_tabla()

    def next_page():
        global page_number
        if page_number * records_per_page < total_records:
            page_number += 1
            llenar_tabla()

    def mostrar_todos():
        global page_number
        page_number = 1
        filtro_id.set("")
        filtro_nombre.set("")
        llenar_tabla()

    # Configuración del marco y los widgets
    marco = tk.LabelFrame(ventana, text="Formulario Ciudad")
    marco.place(x=50, y=50, width=500, height=550)

    # Botones de cerrar y minimizar
    btnCerrar = tk.Button(marco, text="X", command=ventana.destroy, bg="red", fg="white")
    btnCerrar.place(x=470, y=10, width=20, height=20)

    btnMinimizar = tk.Button(marco, text="-", command=lambda: ventana.iconify(), bg="yellow", fg="black")
    btnMinimizar.place(x=450, y=10, width=20, height=20)

    # Campos de entrada para filtrar por ID y Nombre de la ciudad
    tk.Label(marco, text="ID Ciudad").grid(column=0, row=0, padx=5, pady=5)
    txtIdCiudad = tk.Entry(marco, textvariable=filtro_id)
    txtIdCiudad.grid(column=1, row=0)
    txtIdCiudad.bind("<KeyRelease>", lambda event: llenar_tabla())

    tk.Label(marco, text="Buscar por Nombre").grid(column=2, row=0, padx=5, pady=5)
    txtNombreCiudad = tk.Entry(marco, textvariable=filtro_nombre)
    txtNombreCiudad.grid(column=3, row=0)
    txtNombreCiudad.bind("<KeyRelease>", lambda event: llenar_tabla())

    # Campo de entrada original para crear y modificar registros
    tk.Label(marco, text="Nombre Ciudad").grid(column=0, row=1, padx=5, pady=5)
    txtNombre = tk.Entry(marco, textvariable=nombre)
    txtNombre.grid(column=1, row=1)

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=2, columnspan=4)

    tvbancos = ttk.Treeview(marco, selectmode='none')
    tvbancos["columns"] = ("ID Ciudad", "Nombre Ciudad")
    tvbancos.column("#0", width=0, stretch='no')
    tvbancos.column("ID Ciudad", width=150, anchor='center')
    tvbancos.column("Nombre Ciudad", width=150, anchor='center')
    tvbancos.heading("#0", text="")
    tvbancos.heading("ID Ciudad", text="ID Ciudad", anchor='center')
    tvbancos.heading("Nombre Ciudad", text="Nombre Ciudad", anchor='center')
    tvbancos.grid(column=0, row=3, columnspan=4, padx=5)
    tvbancos.bind("<<TreeviewSelect>>", seleccionar)

    btnEliminar = tk.Button(marco, text="Eliminar", command=eliminar)
    btnEliminar.grid(column=1, row=4)

    btnNuevo = tk.Button(marco, text="Guardar", command=nuevo)
    btnNuevo.grid(column=2, row=4)

    btnModificar = tk.Button(marco, text="Modificar", command=actualizar)
    btnModificar.grid(column=3, row=4)

    # Botones de paginación
    btnPrev = tk.Button(marco, text="<< Anterior", command=prev_page)
    btnPrev.grid(column=0, row=5, pady=10, sticky='w')

    btnNext = tk.Button(marco, text="Siguiente >>", command=next_page)
    btnNext.grid(column=3, row=5, pady=10, sticky='e')

    btnMostrarTodos = tk.Button(marco, text="Mostrar Todos", command=mostrar_todos)
    btnMostrarTodos.grid(column=3, row=2, padx=10)

    llenar_tabla()
    
def mostrar_cargo(ventana):
    global db, records_per_page, page_number, total_records
    db = DataBase()
    modificar = False
    records_per_page = 10  # Registros por página
    page_number = 1  # Número de página inicial
    total_records = 0  # Total de registros

    id_cargo = tk.StringVar()
    nombre_cargo = tk.StringVar()
    id_jefe = tk.StringVar()

    filtro_nombre_cargo = tk.StringVar()
    filtro_jefe = tk.StringVar()

    def cargar_nombres_cargo():
        sql = "SELECT DISTINCT nombreCargo FROM cargo"
        db.cursor.execute(sql)
        return [fila[0] for fila in db.cursor.fetchall()]

    def cargar_jefes():
        sql = "SELECT id, nombreJefe FROM jefe"
        db.cursor.execute(sql)
        return db.cursor.fetchall()

    def obtener_nombre_jefe(id_jefe):
        sql = "SELECT nombreJefe FROM jefe WHERE id=%s"
        db.cursor.execute(sql, (id_jefe,))
        result = db.cursor.fetchone()
        return result[0] if result else "Desconocido"

    def obtener_id_jefe(nombre_jefe):
        sql = "SELECT id FROM jefe WHERE nombreJefe=%s"
        db.cursor.execute(sql, (nombre_jefe,))
        result = db.cursor.fetchone()
        return result[0] if result else None

    def seleccionar(event):
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvEstudiantes.item(id, "values")
            id_cargo.set(valores[0])
            nombre_cargo.set(valores[1])
            id_jefe.set(obtener_id_jefe(valores[2]))

    def vaciar_tabla():
        filas = tvEstudiantes.get_children()
        for fila in filas:
            tvEstudiantes.delete(fila)

    def limpiar():
        id_cargo.set("")
        nombre_cargo.set("")
        id_jefe.set("")
        comboJefe.set("")

    def aplicar_filtros():
        llenar_tabla()

    def mostrar_todos():
        # Restablecer los filtros
        filtro_nombre_cargo.set("")
        filtro_jefe.set("")
        comboFiltroNombreCargo.set("")  # Restablecer el valor visual del combobox
        comboFiltroJefe.set("")
        
        # Llenar la tabla sin aplicar filtros
        llenar_tabla()

    def llenar_tabla():
        vaciar_tabla()

        offset = (page_number - 1) * records_per_page

        sql_count = "SELECT COUNT(*) FROM cargo"
        db.cursor.execute(sql_count)
        global total_records
        total_records = db.cursor.fetchone()[0]

        sql = """
        SELECT c.id, c.nombreCargo, j.nombreJefe
        FROM cargo c
        JOIN jefe j ON c.idJefe = j.id
        WHERE 1=1
        """

        filtros = []
        valores = []

        if filtro_nombre_cargo.get():
            sql += " AND c.nombreCargo = %s"
            valores.append(filtro_nombre_cargo.get())

        if filtro_jefe.get():
            sql += " AND j.nombreJefe = %s"
            valores.append(filtro_jefe.get())

        sql += f" LIMIT {records_per_page} OFFSET {offset}"

        db.cursor.execute(sql, tuple(valores))
        filas = db.cursor.fetchall()

        for fila in filas:
            tvEstudiantes.insert("", 'end', fila[0], text=fila[0], values=(fila[0], fila[1], fila[2]))

        actualizar_botones_paginacion()

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tvEstudiantes.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Seleccionar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tvEstudiantes.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def validar():
        return len(id_cargo.get()) > 0 and len(nombre_cargo.get()) > 0 and id_jefe.get() in [str(jefe[0]) for jefe in jefes]

    def eliminar():
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM cargo WHERE id=%s"
                db.cursor.execute(sql, (id,))
                db.connection.commit()
                tvEstudiantes.delete(id)
                lblMensaje.config(text="Se ha eliminado el registro correctamente")
                limpiar()
            else:
                lblMensaje.config(text="Seleccione un registro para eliminar")

    def nuevo():
        if not modificar:
            if validar():
                val = (id_cargo.get(), nombre_cargo.get(), id_jefe.get())
                sql = "INSERT INTO cargo (id, nombreCargo, idJefe) VALUES (%s, %s, %s)"
                db.cursor.execute(sql, val)
                db.connection.commit()
                lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                llenar_tabla()
                limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos o ID de jefe no válido", fg="red")
        else:
            modificarFalse()

    def actualizar():
        if modificar:
            if validar():
                seleccion = tvEstudiantes.selection()
                if seleccion:
                    id = seleccion[0]
                    val = (nombre_cargo.get(), id_jefe.get())
                    sql = "UPDATE cargo SET nombreCargo=%s, idJefe=%s WHERE id=%s"
                    db.cursor.execute(sql, val + (id,))
                    db.connection.commit()
                    lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                    llenar_tabla()
                    limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos o ID de jefe no válido", fg="red")
        else:
            modificarTrue()

    def ir_a_pagina_anterior():
        global page_number
        if page_number > 1:
            page_number -= 1
            llenar_tabla()

    def ir_a_pagina_siguiente():
        global page_number
        if page_number * records_per_page < total_records:
            page_number += 1
            llenar_tabla()

    def actualizar_botones_paginacion():
        btnAnterior.config(state=tk.NORMAL if page_number > 1 else tk.DISABLED)
        btnSiguiente.config(state=tk.NORMAL if page_number * records_per_page < total_records else tk.DISABLED)

    def close_window():
        ventana.destroy()

    def minimize_window():
        ventana.iconify()

    marco = tk.LabelFrame(ventana, text="Formulario Cargo")
    marco.place(x=50, y=50, width=1600, height=1400)

    btnCerrar = tk.Button(marco, text="X", command=close_window, bg="red", fg="white")
    btnCerrar.place(x=570, y=10, width=20, height=20)

    btnMinimizar = tk.Button(marco, text="-", command=minimize_window, bg="yellow", fg="black")
    btnMinimizar.place(x=550, y=10, width=20, height=20)

    tk.Label(marco, text="ID Cargo").grid(column=0, row=0, padx=5, pady=5)
    txtIdCargo = tk.Entry(marco, textvariable=id_cargo)
    txtIdCargo.grid(column=1, row=0)

    tk.Label(marco, text="Nombre Cargo").grid(column=0, row=1, padx=5, pady=5)
    txtNombreCargo = tk.Entry(marco, textvariable=nombre_cargo)
    txtNombreCargo.grid(column=1, row=1)

    tk.Label(marco, text="Jefe").grid(column=0, row=2, padx=5, pady=5)
    jefes = cargar_jefes()
    jefes_dict = {jefe[1]: jefe[0] for jefe in jefes}

    comboJefe = ttk.Combobox(marco, values=[jefe[1] for jefe in jefes])
    comboJefe.grid(column=1, row=2)

    def on_combobox_select(event):
        selected_name = comboJefe.get()
        id_jefe.set(jefes_dict.get(selected_name, ""))

    comboJefe.bind("<<ComboboxSelected>>", on_combobox_select)

    # Agregar Combobox de filtros
    tk.Label(marco, text="Filtrar por Nombre Cargo").grid(column=2, row=0, padx=5, pady=5)
    comboFiltroNombreCargo = ttk.Combobox(marco, values=cargar_nombres_cargo(), textvariable=filtro_nombre_cargo)
    comboFiltroNombreCargo.grid(column=3, row=0, padx=5, pady=5)
    comboFiltroNombreCargo.bind("<<ComboboxSelected>>", lambda event: aplicar_filtros())

    tk.Label(marco, text="Filtrar por Jefe").grid(column=2, row=1, padx=5, pady=5)
    comboFiltroJefe = ttk.Combobox(marco, values=[jefe[1] for jefe in jefes], textvariable=filtro_jefe)
    comboFiltroJefe.grid(column=3, row=1, padx=5, pady=5)
    comboFiltroJefe.bind("<<ComboboxSelected>>", lambda event: aplicar_filtros())

    # Botón para mostrar todos los registros
    btnMostrarTodos = tk.Button(marco, text="Mostrar Todos", command=mostrar_todos)
    btnMostrarTodos.grid(column=2, row=3, padx=5, pady=5)

    btnNuevo = tk.Button(marco, text="Guardar", command=nuevo)
    btnNuevo.grid(column=0, row=3, padx=5, pady=5)

    btnModificar = tk.Button(marco, text="Seleccionar", command=actualizar)
    btnModificar.grid(column=1, row=3, padx=5, pady=5)

    btnEliminar = tk.Button(marco, text="Eliminar", command=eliminar)
    btnEliminar.grid(column=2, row=3, padx=5, pady=5)
    btnEliminar.config(state='disabled')

    lblMensaje = tk.Label(marco, text="")
    lblMensaje.grid(column=0, row=4, columnspan=4)

    tvEstudiantes = ttk.Treeview(marco, selectmode='none')
    tvEstudiantes.grid(column=0, row=5, columnspan=4, padx=5, pady=5)
    tvEstudiantes['columns'] = ('ID', 'Nombre Cargo', 'Jefe')

    tvEstudiantes.column('#0', width=0, stretch=tk.NO)
    tvEstudiantes.column('ID', anchor=tk.W, width=100)
    tvEstudiantes.column('Nombre Cargo', anchor=tk.W, width=200)
    tvEstudiantes.column('Jefe', anchor=tk.W, width=200)

    tvEstudiantes.heading('#0', text='', anchor=tk.W)
    tvEstudiantes.heading('ID', text='ID', anchor=tk.W)
    tvEstudiantes.heading('Nombre Cargo', text='Nombre Cargo', anchor=tk.W)
    tvEstudiantes.heading('Jefe', text='Jefe', anchor=tk.W)

    tvEstudiantes.bind("<<TreeviewSelect>>", seleccionar)

    # Botones de paginación
    btnAnterior = tk.Button(marco, text="Anterior", command=ir_a_pagina_anterior)
    btnAnterior.grid(column=0, row=6, padx=5, pady=5)

    btnSiguiente = tk.Button(marco, text="Siguiente", command=ir_a_pagina_siguiente)
    btnSiguiente.grid(column=1, row=6, padx=5, pady=5)

    llenar_tabla()
    
def mostrar_departamento(ventana):
    # Inicializar variables
    db = DataBase()
    modificar = False
    id_departamento = tk.StringVar()
    nombre_departamento = tk.StringVar()

    def seleccionar(event):
        seleccion = tvDepartamentos.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvDepartamentos.item(id, "values")
            id_departamento.set(valores[0])
            nombre_departamento.set(valores[1])

    def vaciar_tabla():
        filas = tvDepartamentos.get_children()
        for fila in filas:
            tvDepartamentos.delete(fila)

    def limpiar():
        id_departamento.set("")
        nombre_departamento.set("")

    def llenar_tabla():
        vaciar_tabla()
        sql = "SELECT id, nombreDepartamento FROM departamento"
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        for fila in filas:
            id = fila[0]
            tvDepartamentos.insert("", 'end', id, text=id, values=(fila[0], fila[1]))

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tvDepartamentos.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Seleccionar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tvDepartamentos.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def validar():
        return len(id_departamento.get()) > 0 and len(nombre_departamento.get()) > 0

    def eliminar():
        seleccion = tvDepartamentos.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM departamento WHERE id=%s"
                db.cursor.execute(sql, (id,))
                db.connection.commit()
                tvDepartamentos.delete(id)
                lblMensaje.config(text="Se ha eliminado el registro correctamente")
                limpiar()
            else:
                lblMensaje.config(text="Seleccione un registro para eliminar")

    def nuevo():
        if not modificar:
            if validar():
                val = (id_departamento.get(), nombre_departamento.get())
                sql = "INSERT INTO departamento (id, nombreDepartamento) VALUES (%s, %s)"
                db.cursor.execute(sql, val)
                db.connection.commit()
                lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                llenar_tabla()
                limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarFalse()

    def actualizar():
        if modificar:
            if validar():
                seleccion = tvDepartamentos.selection()
                if seleccion:
                    id = seleccion[0]
                    val = (nombre_departamento.get(),)
                    sql = "UPDATE departamento SET nombreDepartamento=%s WHERE id=%s"
                    db.cursor.execute(sql, val + (id,))
                    db.connection.commit()
                    lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                    llenar_tabla()
                    limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarTrue()

    def close_window():
        ventana.destroy()

    def minimize_window():
        ventana.iconify()

    # Configuración del marco y los widgets
    marco = tk.LabelFrame(ventana, text="Formulario Departamento")
    marco.place(x=50, y=50, width=500, height=400)

    # Botones de cerrar y minimizar
    btnCerrar = tk.Button(marco, text="X", command=close_window, bg="red", fg="white")
    btnCerrar.place(x=470, y=10, width=20, height=20)
    
    btnMinimizar = tk.Button(marco, text="-", command=minimize_window, bg="yellow", fg="black")
    btnMinimizar.place(x=450, y=10, width=20, height=20)

    tk.Label(marco, text="ID").grid(column=0, row=0, padx=5, pady=5)
    txtIdDepartamento = tk.Entry(marco, textvariable=id_departamento)
    txtIdDepartamento.grid(column=1, row=0)

    tk.Label(marco, text="Nombre").grid(column=0, row=1, padx=5, pady=5)
    txtNombreDepartamento = tk.Entry(marco, textvariable=nombre_departamento)
    txtNombreDepartamento.grid(column=1, row=1)

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=2, columnspan=4)

    tvDepartamentos = ttk.Treeview(marco, selectmode='none')
    tvDepartamentos["columns"] = ("ID", "Nombre")
    tvDepartamentos.column("#0", width=0, stretch='no')
    tvDepartamentos.column("ID", width=150, anchor='center')
    tvDepartamentos.column("Nombre", width=150, anchor='center')
    tvDepartamentos.heading("#0", text="")
    tvDepartamentos.heading("ID", text="ID", anchor='center')
    tvDepartamentos.heading("Nombre", text="Nombre", anchor='center')
    tvDepartamentos.grid(column=0, row=3, columnspan=4, padx=5)
    tvDepartamentos.bind("<<TreeviewSelect>>", seleccionar)

    btnEliminar = tk.Button(marco, text="Eliminar", command=lambda: eliminar())
    btnEliminar.grid(column=1, row=4)
    btnNuevo = tk.Button(marco, text="Guardar", command=lambda: nuevo())
    btnNuevo.grid(column=2, row=4)
    btnModificar = tk.Button(marco, text="Seleccionar", command=lambda: actualizar())
    btnModificar.grid(column=3, row=4)

    llenar_tabla()

def mostrar_rh(ventana):
    # Inicializar variables
    db = DataBase()
    modificar = False
    id_rh = tk.StringVar()
    tipo_rh = tk.StringVar()

    def seleccionar(event):
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvEstudiantes.item(id, "values")
            id_rh.set(valores[0])
            tipo_rh.set(valores[1])

    marco = tk.LabelFrame(ventana, text="Formulario RH")
    marco.place(x=50, y=50, width=500, height=400)

    tk.Label(marco, text="ID").grid(column=0, row=0, padx=5, pady=5)
    txtIdRh = tk.Entry(marco, textvariable=id_rh)
    txtIdRh.grid(column=1, row=0)

    tk.Label(marco, text="Tipo RH").grid(column=0, row=1, padx=5, pady=5)
    txtTipoRh = tk.Entry(marco, textvariable=tipo_rh)
    txtTipoRh.grid(column=1, row=1)

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=2, columnspan=4)

    tvEstudiantes = ttk.Treeview(marco, selectmode='none')
    tvEstudiantes["columns"] = ("ID", "Tipo RH")
    tvEstudiantes.column("#0", width=0, stretch='no')
    tvEstudiantes.column("ID", width=150, anchor='center')
    tvEstudiantes.column("Tipo RH", width=150, anchor='center')
    tvEstudiantes.heading("#0", text="")
    tvEstudiantes.heading("ID", text="ID", anchor='center')
    tvEstudiantes.heading("Tipo RH", text="Tipo RH", anchor='center')
    tvEstudiantes.grid(column=0, row=3, columnspan=4, padx=5)
    tvEstudiantes.bind("<<TreeviewSelect>>", seleccionar)

    btnEliminar = tk.Button(marco, text="Eliminar", command=lambda: eliminar())
    btnEliminar.grid(column=1, row=4)
    btnNuevo = tk.Button(marco, text="Guardar", command=lambda: nuevo())
    btnNuevo.grid(column=2, row=4)
    btnModificar = tk.Button(marco, text="Seleccionar", command=lambda: actualizar())
    btnModificar.grid(column=3, row=4)

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tvEstudiantes.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Seleccionar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tvEstudiantes.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def validar():
        return len(id_rh.get()) > 0 and len(tipo_rh.get()) > 0

    def limpiar():
        id_rh.set("")
        tipo_rh.set("")

    def vaciar_tabla():
        filas = tvEstudiantes.get_children()
        for fila in filas:
            tvEstudiantes.delete(fila)

    def llenar_tabla():
        vaciar_tabla()
        sql = "SELECT id, tipoRh FROM rh"
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        for fila in filas:
            id = fila[0]
            tvEstudiantes.insert("", 'end', id, text=id, values=(fila[0], fila[1]))

    def eliminar():
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM rh WHERE id=%s"
                db.cursor.execute(sql, (id,))
                db.connection.commit()
                tvEstudiantes.delete(id)
                lblMensaje.config(text="Se ha eliminado el registro correctamente")
                limpiar()
            else:
                lblMensaje.config(text="Seleccione un registro para eliminar")

    def nuevo():
        if not modificar:
            if validar():
                val = (id_rh.get(), tipo_rh.get())
                sql = "INSERT INTO rh (id, tipoRh) VALUES (%s, %s)"
                db.cursor.execute(sql, val)
                db.connection.commit()
                lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                llenar_tabla()
                limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarFalse()

    def actualizar():
        if modificar:
            if validar():
                seleccion = tvEstudiantes.selection()
                if seleccion:
                    id = seleccion[0]
                    val = (tipo_rh.get(),)
                    sql = "UPDATE rh SET tipoRh=%s WHERE id=%s"
                    db.cursor.execute(sql, val + (id,))
                    db.connection.commit()
                    lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                    llenar_tabla()
                    limpiar()
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarTrue()

    llenar_tabla()

def mostrar_departamento(ventana):
    # Inicializar variables
    db = DataBase()
    modificar = False
    id_departamento = tk.StringVar()
    nombre_departamento = tk.StringVar()

    def seleccionar(event):
        seleccion = tvDepartamentos.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvDepartamentos.item(id, "values")
            id_departamento.set(valores[0])
            nombre_departamento.set(valores[1])

    def vaciar_tabla():
        filas = tvDepartamentos.get_children()
        for fila in filas:
            tvDepartamentos.delete(fila)

    def limpiar():
        id_departamento.set("")
        nombre_departamento.set("")

    def llenar_tabla():
        vaciar_tabla()
        sql = "SELECT id, nombreDepartamento FROM departamentos"  # Cambiado el nombre de la tabla a 'departamentos'
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        for fila in filas:
            id = fila[0]
            tvDepartamentos.insert("", 'end', id, text=id, values=(fila[0], fila[1]))

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tvDepartamentos.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Seleccionar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tvDepartamentos.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def validar():
        return len(id_departamento.get()) > 0 and len(nombre_departamento.get()) > 0

    def eliminar():
        seleccion = tvDepartamentos.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM departamentos WHERE id=%s"  # Cambiado el nombre de la tabla a 'departamentos'
                db.cursor.execute(sql, (id,))
                db.connection.commit()
                tvDepartamentos.delete(id)
                lblMensaje.config(text="Se ha eliminado el registro correctamente", fg="green")
                limpiar()
                llenar_tabla()  # Asegúrate de volver a llenar la tabla después de eliminar
            else:
                lblMensaje.config(text="Seleccione un registro para eliminar", fg="red")

    def nuevo():
        if not modificar:
            if validar():
                val = (id_departamento.get(), nombre_departamento.get())
                sql = "INSERT INTO departamentos (id, nombreDepartamento) VALUES (%s, %s)"  # Cambiado el nombre de la tabla a 'departamentos'
                try:
                    db.cursor.execute(sql, val)
                    db.connection.commit()
                    lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                    llenar_tabla()
                    limpiar()
                except pymysql.err.IntegrityError as e:
                    lblMensaje.config(text=f"Error: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarFalse()

    def actualizar():
        if modificar:
            if validar():
                seleccion = tvDepartamentos.selection()
                if seleccion:
                    id = seleccion[0]
                    val = (nombre_departamento.get(), id)
                    sql = "UPDATE departamentos SET nombreDepartamento=%s WHERE id=%s"  # Cambiado el nombre de la tabla a 'departamentos'
                    try:
                        db.cursor.execute(sql, val)
                        db.connection.commit()
                        lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                        llenar_tabla()
                        limpiar()
                    except pymysql.err.IntegrityError as e:
                        lblMensaje.config(text=f"Error: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarTrue()

    def close_window():
        ventana.destroy()

    def minimize_window():
        ventana.iconify()

    # Configuración del marco y los widgets
    marco = tk.LabelFrame(ventana, text="Formulario Departamento")
    marco.place(x=50, y=50, width=500, height=400)

    # Botones de cerrar y minimizar
    btnCerrar = tk.Button(marco, text="X", command=close_window, bg="red", fg="white")
    btnCerrar.place(x=470, y=10, width=20, height=20)
    
    btnMinimizar = tk.Button(marco, text="-", command=minimize_window, bg="yellow", fg="black")
    btnMinimizar.place(x=450, y=10, width=20, height=20)

    tk.Label(marco, text="ID").grid(column=0, row=0, padx=5, pady=5)
    txtIdDepartamento = tk.Entry(marco, textvariable=id_departamento)
    txtIdDepartamento.grid(column=1, row=0)

    tk.Label(marco, text="Nombre").grid(column=0, row=1, padx=5, pady=5)
    txtNombreDepartamento = tk.Entry(marco, textvariable=nombre_departamento)
    txtNombreDepartamento.grid(column=1, row=1)

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=2, columnspan=4)

    tvDepartamentos = ttk.Treeview(marco, selectmode='none')
    tvDepartamentos["columns"] = ("ID", "Nombre")
    tvDepartamentos.column("#0", width=0, stretch='no')
    tvDepartamentos.column("ID", width=150, anchor='center')
    tvDepartamentos.column("Nombre", width=150, anchor='center')
    tvDepartamentos.heading("#0", text="")
    tvDepartamentos.heading("ID", text="ID", anchor='center')
    tvDepartamentos.heading("Nombre", text="Nombre", anchor='center')
    tvDepartamentos.grid(column=0, row=3, columnspan=4, padx=5)
    tvDepartamentos.bind("<<TreeviewSelect>>", seleccionar)

    btnEliminar = tk.Button(marco, text="Eliminar", command=lambda: eliminar())
    btnEliminar.grid(column=1, row=4)
    btnNuevo = tk.Button(marco, text="Guardar", command=lambda: nuevo())
    btnNuevo.grid(column=2, row=4)
    btnModificar = tk.Button(marco, text="Seleccionar", command=lambda: actualizar())
    btnModificar.grid(column=3, row=4)

    llenar_tabla()

def exportar_a_excel_dependencias():
    try:
        sql = "SELECT id, nombreDependencia FROM dependencia"
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        if filas:
            df = pd.DataFrame(filas, columns=['ID', 'Nombre'])

            filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                    filetypes=[("Excel files", "*.xlsx"), 
                                                               ("All files", "*.*")])
            
            if filepath:
                df.to_excel(filepath, index=False)
                messagebox.showinfo("Exportación Exitosa", f"Datos exportados a {filepath}")
                os.startfile(filepath)
            else:
                messagebox.showwarning("Cancelado", "Exportación cancelada por el usuario")
        else:
            messagebox.showwarning("Sin Datos", "No hay datos para exportar")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar los datos: {e}")

def mostrar_dependencia(ventana):
    global db
    db = DataBase()
    modificar = False
    id_dependencia = tk.StringVar()
    nombre_dependencia = tk.StringVar()
    filtro_nombre = tk.StringVar()  # Variable para el filtro de Nombre

    # Variables de paginación
    page_number = 1
    records_per_page = 10

    def seleccionar(event):
        seleccion = tvDependencias.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvDependencias.item(id, "values")
            id_dependencia.set(valores[0])
            nombre_dependencia.set(valores[1])

    def vaciar_tabla():
        filas = tvDependencias.get_children()
        for fila in filas:
            tvDependencias.delete(fila)

    def limpiar():
        id_dependencia.set("")
        nombre_dependencia.set("")
        filtro_nombre.set("")  # Limpiar el filtro también

    def llenar_tabla(filtro=None):
        vaciar_tabla()
        try:
            offset = (page_number - 1) * records_per_page
            if filtro:
                sql = "SELECT id, nombreDependencia FROM dependencia WHERE nombreDependencia=%s LIMIT %s OFFSET %s"
                db.cursor.execute(sql, (filtro, records_per_page, offset))
            else:
                sql = "SELECT id, nombreDependencia FROM dependencia LIMIT %s OFFSET %s"
                db.cursor.execute(sql, (records_per_page, offset))
            filas = db.cursor.fetchall()
            for fila in filas:
                id = fila[0]
                tvDependencias.insert("", 'end', id, text=id, values=(fila[0], fila[1]))
        except Exception as e:
            lblMensaje.config(text=f"Error al llenar la tabla: {e}", fg="red")

    def cargar_dependencias_en_combobox():
        try:
            sql = "SELECT DISTINCT nombreDependencia FROM dependencia"
            db.cursor.execute(sql)
            dependencias = db.cursor.fetchall()
            nombres = [fila[0] for fila in dependencias]  # Obtener los nombres de las dependencias
            cbFiltroNombre['values'] = nombres  # Actualizar el Combobox con los nombres
        except Exception as e:
            lblMensaje.config(text=f"Error al cargar nombres: {e}", fg="red")

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tvDependencias.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Seleccionar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tvDependencias.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def validar():
        return len(id_dependencia.get()) > 0 and len(nombre_dependencia.get()) > 0

    def eliminar():
        seleccion = tvDependencias.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM dependencia WHERE id=%s"
                db.cursor.execute(sql, (id,))
                db.connection.commit()
                tvDependencias.delete(id)
                lblMensaje.config(text="Se ha eliminado el registro correctamente", fg="green")
                limpiar()
                llenar_tabla()
            else:
                lblMensaje.config(text="Seleccione un registro para eliminar", fg="red")

    def nuevo():
        if not modificar:
            if validar():
                val = (id_dependencia.get(), nombre_dependencia.get())
                sql = "INSERT INTO dependencia (id, nombreDependencia) VALUES (%s, %s)"
                try:
                    db.cursor.execute(sql, val)
                    db.connection.commit()
                    lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                    llenar_tabla()
                    limpiar()
                    cargar_dependencias_en_combobox()  # Actualizar el Combobox después de agregar una nueva dependencia
                except pymysql.err.IntegrityError as e:
                    lblMensaje.config(text=f"Error: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarFalse()

    def actualizar():
        if modificar:
            if validar():
                seleccion = tvDependencias.selection()
                if seleccion:
                    id = seleccion[0]
                    val = (nombre_dependencia.get(),)
                    sql = "UPDATE dependencia SET nombreDependencia=%s WHERE id=%s"
                    try:
                        db.cursor.execute(sql, val + (id,))
                        db.connection.commit()
                        lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                        llenar_tabla()
                        limpiar()
                        cargar_dependencias_en_combobox()  # Actualizar el Combobox después de la actualización
                    except pymysql.err.IntegrityError as e:
                        lblMensaje.config(text=f"Error: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarTrue()

    def close_window():
        ventana.destroy()

    def minimize_window():
        ventana.iconify()

    def ir_a_pagina_anterior():
        nonlocal page_number
        if page_number > 1:
            page_number -= 1
            llenar_tabla()
            actualizar_botones_paginacion()

    def ir_a_pagina_siguiente():
        nonlocal page_number
        page_number += 1
        llenar_tabla()
        actualizar_botones_paginacion()

    def actualizar_botones_paginacion():
        # Actualizar la disponibilidad de los botones
        btnAnterior.config(state='normal' if page_number > 1 else 'disabled')
        btnSiguiente.config(state='normal')  # Aquí puedes añadir lógica para deshabilitar si no hay más registros

    def aplicar_filtro(event):
        filtro = filtro_nombre.get()
        if filtro:
            llenar_tabla(filtro)
        limpiar()

    # Configuración del marco y los widgets
    marco = tk.LabelFrame(ventana, text="Formulario Dependencia")
    marco.place(x=50, y=50, width=500, height=450)

    # Botones de cerrar y minimizar
    btnCerrar = tk.Button(marco, text="X", command=close_window, bg="red", fg="white")
    btnCerrar.place(x=470, y=10, width=20, height=20)

    btnMinimizar = tk.Button(marco, text="-", command=minimize_window, bg="yellow", fg="black")
    btnMinimizar.place(x=450, y=10, width=20, height=20)

    tk.Label(marco, text="ID").grid(column=0, row=0, padx=5, pady=5)
    txtIdDependencia = tk.Entry(marco, textvariable=id_dependencia)
    txtIdDependencia.grid(column=1, row=0)

    tk.Label(marco, text="Nombre").grid(column=0, row=1, padx=5, pady=5)
    txtNombreDependencia = tk.Entry(marco, textvariable=nombre_dependencia)
    txtNombreDependencia.grid(column=1, row=1)

    # Combobox para filtro por Nombre
    tk.Label(marco, text="Filtrar por Nombre").grid(column=0, row=2, padx=5, pady=5)
    cbFiltroNombre = ttk.Combobox(marco, textvariable=filtro_nombre)
    cbFiltroNombre.grid(column=1, row=2)
    cbFiltroNombre.bind("<<ComboboxSelected>>", aplicar_filtro)

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=3, columnspan=4)

    tvDependencias = ttk.Treeview(marco, selectmode='none')
    tvDependencias["columns"] = ("ID", "Nombre")
    tvDependencias.column("#0", width=0, stretch='no')
    tvDependencias.column("ID", width=150, anchor='center')
    tvDependencias.column("Nombre", width=150, anchor='center')
    tvDependencias.heading("#0", text="")
    tvDependencias.heading("ID", text="ID", anchor='center')
    tvDependencias.heading("Nombre", text="Nombre", anchor='center')
    tvDependencias.grid(column=0, row=4, columnspan=4, padx=5)
    tvDependencias.bind("<<TreeviewSelect>>", seleccionar)

    btnEliminar = tk.Button(marco, text="Eliminar", command=lambda: eliminar())
    btnEliminar.grid(column=1, row=5)

    btnNuevo = tk.Button(marco, text="Guardar", command=lambda: nuevo())
    btnNuevo.grid(column=2, row=5)

    btnModificar = tk.Button(marco, text="Seleccionar", command=lambda: actualizar())
    btnModificar.grid(column=3, row=5)

    # Botones de Paginación
    btnAnterior = tk.Button(marco, text="<< Anterior", command=ir_a_pagina_anterior)
    btnAnterior.grid(column=0, row=6)

    btnSiguiente = tk.Button(marco, text="Siguiente >>", command=ir_a_pagina_siguiente)
    btnSiguiente.grid(column=1, row=6)

    # Añadir botón para exportar a Excel
    btnExportar = tk.Button(marco, text="Exportar a Excel", command=exportar_a_excel_dependencias)
    btnExportar.grid(column=2, row=6, pady=10)

    llenar_tabla()
    cargar_dependencias_en_combobox()  # Llenar el Combobox con los nombres de las dependencias al iniciar
    actualizar_botones_paginacion()

def mostrar_rh(ventana):
    # Inicializar variables
    db = DataBase()
    modificar = False
    id_rh = tk.StringVar()
    tipo_rh = tk.StringVar()

    def seleccionar(event):
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvEstudiantes.item(id, "values")
            id_rh.set(valores[0])
            tipo_rh.set(valores[1])

    marco = tk.LabelFrame(ventana, text="Formulario RH")
    marco.place(x=50, y=50, width=500, height=400)

    # Botones de cerrar y minimizar
    btnCerrar = tk.Button(marco, text="X", command=ventana.destroy, bg="red", fg="white")
    btnCerrar.place(x=470, y=10, width=20, height=20)
    
    btnMinimizar = tk.Button(marco, text="-", command=lambda: ventana.iconify(), bg="yellow", fg="black")
    btnMinimizar.place(x=450, y=10, width=20, height=20)

    tk.Label(marco, text="ID").grid(column=0, row=0, padx=5, pady=5)
    txtIdRh = tk.Entry(marco, textvariable=id_rh)
    txtIdRh.grid(column=1, row=0)

    tk.Label(marco, text="Tipo RH").grid(column=0, row=1, padx=5, pady=5)
    txtTipoRh = tk.Entry(marco, textvariable=tipo_rh)
    txtTipoRh.grid(column=1, row=1)

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=2, columnspan=4)

    tvEstudiantes = ttk.Treeview(marco, selectmode='none')
    tvEstudiantes["columns"] = ("ID", "Tipo RH")
    tvEstudiantes.column("#0", width=0, stretch='no')
    tvEstudiantes.column("ID", width=150, anchor='center')
    tvEstudiantes.column("Tipo RH", width=150, anchor='center')
    tvEstudiantes.heading("#0", text="")
    tvEstudiantes.heading("ID", text="ID", anchor='center')
    tvEstudiantes.heading("Tipo RH", text="Tipo RH", anchor='center')
    tvEstudiantes.grid(column=0, row=3, columnspan=4, padx=5)
    tvEstudiantes.bind("<<TreeviewSelect>>", seleccionar)

    btnEliminar = tk.Button(marco, text="Eliminar", command=lambda: eliminar())
    btnEliminar.grid(column=1, row=4)
    btnNuevo = tk.Button(marco, text="Guardar", command=lambda: nuevo())
    btnNuevo.grid(column=2, row=4)
    btnModificar = tk.Button(marco, text="Seleccionar", command=lambda: actualizar())
    btnModificar.grid(column=3, row=4)

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tvEstudiantes.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Seleccionar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tvEstudiantes.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def validar():
        return len(id_rh.get()) > 0 and len(tipo_rh.get()) > 0

    def limpiar():
        id_rh.set("")
        tipo_rh.set("")

    def vaciar_tabla():
        filas = tvEstudiantes.get_children()
        for fila in filas:
            tvEstudiantes.delete(fila)

    def llenar_tabla():
        vaciar_tabla()
        sql = "SELECT id, tipoRh FROM rh"
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        for fila in filas:
            id = fila[0]
            tvEstudiantes.insert("", 'end', id, text=id, values=(fila[0], fila[1]))

    def eliminar():
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM rh WHERE id=%s"
                db.cursor.execute(sql, (id,))
                db.connection.commit()
                tvEstudiantes.delete(id)
                lblMensaje.config(text="Se ha eliminado el registro correctamente", fg="green")
                limpiar()
                llenar_tabla()  # Asegúrate de volver a llenar la tabla después de eliminar
            else:
                lblMensaje.config(text="Seleccione un registro para eliminar", fg="red")

    def nuevo():
        if not modificar:
            if validar():
                val = (id_rh.get(), tipo_rh.get())
                sql = "INSERT INTO rh (id, tipoRh) VALUES (%s, %s)"
                try:
                    db.cursor.execute(sql, val)
                    db.connection.commit()
                    lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                    llenar_tabla()
                    limpiar()
                except pymysql.err.IntegrityError as e:
                    lblMensaje.config(text=f"Error: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarFalse()

    def actualizar():
        if modificar:
            if validar():
                seleccion = tvEstudiantes.selection()
                if seleccion:
                    id = seleccion[0]
                    val = (tipo_rh.get(),)
                    sql = "UPDATE rh SET tipoRh=%s WHERE id=%s"
                    try:
                        db.cursor.execute(sql, val + (id,))
                        db.connection.commit()
                        lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                        llenar_tabla()
                        limpiar()
                    except pymysql.err.IntegrityError as e:
                        lblMensaje.config(text=f"Error: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarTrue()

    llenar_tabla()

def mostrar_genero(ventana):
    # Inicializar variables
    db = DataBase()
    modificar = False
    id_genero = tk.StringVar()
    nombre_genero = tk.StringVar()

    def seleccionar(event):
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvEstudiantes.item(id, "values")
            id_genero.set(valores[0])
            nombre_genero.set(valores[1])

    marco = tk.LabelFrame(ventana, text="Formulario Género")
    marco.place(x=50, y=50, width=500, height=400)

    # Botones para cerrar y minimizar
    btnCerrar = tk.Button(marco, text="X", command=ventana.destroy, bg="red", fg="white")
    btnCerrar.place(x=470, y=10, width=20, height=20)
    
    btnMinimizar = tk.Button(marco, text="-", command=lambda: ventana.iconify(), bg="yellow", fg="black")
    btnMinimizar.place(x=450, y=10, width=20, height=20)

    tk.Label(marco, text="ID").grid(column=0, row=0, padx=5, pady=5)
    txtIdGenero = tk.Entry(marco, textvariable=id_genero)
    txtIdGenero.grid(column=1, row=0)

    tk.Label(marco, text="Nombre Género").grid(column=0, row=1, padx=5, pady=5)
    txtNombreGenero = tk.Entry(marco, textvariable=nombre_genero)
    txtNombreGenero.grid(column=1, row=1)

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=2, columnspan=4)

    tvEstudiantes = ttk.Treeview(marco, selectmode='none')
    tvEstudiantes["columns"] = ("ID", "Nombre Género")
    tvEstudiantes.column("#0", width=0, stretch='no')
    tvEstudiantes.column("ID", width=150, anchor='center')
    tvEstudiantes.column("Nombre Género", width=150, anchor='center')
    tvEstudiantes.heading("#0", text="")
    tvEstudiantes.heading("ID", text="ID", anchor='center')
    tvEstudiantes.heading("Nombre Género", text="Nombre Género", anchor='center')
    tvEstudiantes.grid(column=0, row=3, columnspan=4, padx=5)
    tvEstudiantes.bind("<<TreeviewSelect>>", seleccionar)

    btnEliminar = tk.Button(marco, text="Eliminar", command=lambda: eliminar())
    btnEliminar.grid(column=1, row=4)
    btnNuevo = tk.Button(marco, text="Guardar", command=lambda: nuevo())
    btnNuevo.grid(column=2, row=4)
    btnModificar = tk.Button(marco, text="Seleccionar", command=lambda: actualizar())
    btnModificar.grid(column=3, row=4)

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tvEstudiantes.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Seleccionar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tvEstudiantes.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def validar():
        return len(id_genero.get()) > 0 and len(nombre_genero.get()) > 0

    def limpiar():
        id_genero.set("")
        nombre_genero.set("")

    def vaciar_tabla():
        filas = tvEstudiantes.get_children()
        for fila in filas:
            tvEstudiantes.delete(fila)

    def llenar_tabla():
        vaciar_tabla()
        sql = "SELECT id, nombreGenero FROM genero"
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        for fila in filas:
            id = fila[0]
            tvEstudiantes.insert("", 'end', id, text=id, values=(fila[0], fila[1]))

    def eliminar():
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM genero WHERE id=%s"
                db.cursor.execute(sql, (id,))
                db.connection.commit()
                tvEstudiantes.delete(id)
                lblMensaje.config(text="Se ha eliminado el registro correctamente", fg="green")
                limpiar()
                llenar_tabla()  # Asegúrate de volver a llenar la tabla después de eliminar
            else:
                lblMensaje.config(text="Seleccione un registro para eliminar", fg="red")

    def nuevo():
        if not modificar:
            if validar():
                val = (id_genero.get(), nombre_genero.get())
                sql = "INSERT INTO genero (id, nombreGenero) VALUES (%s, %s)"
                try:
                    db.cursor.execute(sql, val)
                    db.connection.commit()
                    lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                    llenar_tabla()
                    limpiar()
                except pymysql.err.IntegrityError as e:
                    lblMensaje.config(text=f"Error: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarFalse()

    def actualizar():
        if modificar:
            if validar():
                seleccion = tvEstudiantes.selection()
                if seleccion:
                    id = seleccion[0]
                    val = (nombre_genero.get(),)
                    sql = "UPDATE genero SET nombreGenero=%s WHERE id=%s"
                    try:
                        db.cursor.execute(sql, val + (id,))
                        db.connection.commit()
                        lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                        llenar_tabla()
                        limpiar()
                    except pymysql.err.IntegrityError as e:
                        lblMensaje.config(text=f"Error: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarTrue()

    llenar_tabla()

def mostrar_tipodedocumento(ventana):
    # Inicializar variables
    db = DataBase()
    modificar = False
    id_documento = tk.StringVar()
    descripcion_documento = tk.StringVar()

    def seleccionar(event):
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            valores = tvEstudiantes.item(id, "values")
            id_documento.set(valores[0])
            descripcion_documento.set(valores[1])

    marco = tk.LabelFrame(ventana, text="Formulario Tipo de Documento")
    marco.place(x=50, y=50, width=500, height=400)

    # Botones para cerrar y minimizar
    btnCerrar = tk.Button(marco, text="X", command=ventana.destroy, bg="red", fg="white")
    btnCerrar.place(x=470, y=10, width=20, height=20)
    
    btnMinimizar = tk.Button(marco, text="-", command=lambda: ventana.iconify(), bg="yellow", fg="black")
    btnMinimizar.place(x=450, y=10, width=20, height=20)

    tk.Label(marco, text="ID").grid(column=0, row=0, padx=5, pady=5)
    txtIdDocumento = tk.Entry(marco, textvariable=id_documento)
    txtIdDocumento.grid(column=1, row=0)

    tk.Label(marco, text="Descripción Documento").grid(column=0, row=1, padx=5, pady=5)
    txtDescripcionDocumento = tk.Entry(marco, textvariable=descripcion_documento)
    txtDescripcionDocumento.grid(column=1, row=1)

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=2, columnspan=4)

    tvEstudiantes = ttk.Treeview(marco, selectmode='none')
    tvEstudiantes["columns"] = ("ID", "Descripción Documento")
    tvEstudiantes.column("#0", width=0, stretch='no')
    tvEstudiantes.column("ID", width=150, anchor='center')
    tvEstudiantes.column("Descripción Documento", width=300, anchor='center')
    tvEstudiantes.heading("#0", text="")
    tvEstudiantes.heading("ID", text="ID", anchor='center')
    tvEstudiantes.heading("Descripción Documento", text="Descripción Documento", anchor='center')
    tvEstudiantes.grid(column=0, row=3, columnspan=4, padx=5)
    tvEstudiantes.bind("<<TreeviewSelect>>", seleccionar)

    btnEliminar = tk.Button(marco, text="Eliminar", command=lambda: eliminar())
    btnEliminar.grid(column=1, row=4)
    btnNuevo = tk.Button(marco, text="Guardar", command=lambda: nuevo())
    btnNuevo.grid(column=2, row=4)
    btnModificar = tk.Button(marco, text="Seleccionar", command=lambda: actualizar())
    btnModificar.grid(column=3, row=4)

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tvEstudiantes.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Seleccionar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tvEstudiantes.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def validar():
        return len(id_documento.get()) > 0 and len(descripcion_documento.get()) > 0

    def limpiar():
        id_documento.set("")
        descripcion_documento.set("")

    def vaciar_tabla():
        filas = tvEstudiantes.get_children()
        for fila in filas:
            tvEstudiantes.delete(fila)

    def llenar_tabla():
        vaciar_tabla()
        sql = "SELECT id, DescripcionDocumento FROM tipodedocumento"
        db.cursor.execute(sql)
        filas = db.cursor.fetchall()
        for fila in filas:
            id = fila[0]
            tvEstudiantes.insert("", 'end', id, text=id, values=(fila[0], fila[1]))

    def eliminar():
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM tipodedocumento WHERE id=%s"
                db.cursor.execute(sql, (id,))
                db.connection.commit()
                tvEstudiantes.delete(id)
                lblMensaje.config(text="Se ha eliminado el registro correctamente", fg="green")
                limpiar()
                llenar_tabla()  # Asegúrate de volver a llenar la tabla después de eliminar
            else:
                lblMensaje.config(text="Seleccione un registro para eliminar", fg="red")

    def nuevo():
        if not modificar:
            if validar():
                val = (id_documento.get(), descripcion_documento.get())
                sql = "INSERT INTO tipodedocumento (id, DescripcionDocumento) VALUES (%s, %s)"
                try:
                    db.cursor.execute(sql, val)
                    db.connection.commit()
                    lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                    llenar_tabla()
                    limpiar()
                except pymysql.err.IntegrityError as e:
                    lblMensaje.config(text=f"Error: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarFalse()

    def actualizar():
        if modificar:
            if validar():
                seleccion = tvEstudiantes.selection()
                if seleccion:
                    id = seleccion[0]
                    val = (descripcion_documento.get(),)
                    sql = "UPDATE tipodedocumento SET DescripcionDocumento=%s WHERE id=%s"
                    try:
                        db.cursor.execute(sql, val + (id,))
                        db.connection.commit()
                        lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                        llenar_tabla()
                        limpiar()
                    except pymysql.err.IntegrityError as e:
                        lblMensaje.config(text=f"Error: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarTrue()

    llenar_tabla()

class DataBase:
    def __init__(self):
        self.connection = pymysql.connect(host='localhost', user='root', password='', db='contrataciones')
        self.cursor = self.connection.cursor()

def mostrar_tipo_cuenta(ventana):
    # Inicializar variables
    db = DataBase()
    global page_number, records_per_page, total_records
    page_number = 1
    records_per_page = 10
    total_records = 0

    modificar = False
    id_tipo_cuenta = tk.StringVar()
    nombre_tipo_cuenta = tk.StringVar()
    filtro_nombre = tk.StringVar()

    def aplicar_filtro(*args):
        llenar_tabla()

    def seleccionar(event):
        # Actualizar el campo de texto con el registro seleccionado
        if modificar:
            seleccion = tvEstudiantes.selection()
            if seleccion:
                id = seleccion[0]
                valores = tvEstudiantes.item(id, "values")
                id_tipo_cuenta.set(valores[0])
                nombre_tipo_cuenta.set(valores[1])

    def eliminar():
        seleccion = tvEstudiantes.selection()
        if seleccion:
            id = seleccion[0]
            if int(id) > 0:
                sql = "DELETE FROM tipodecuenta WHERE id=%s"
                try:
                    db.cursor.execute(sql, (id,))
                    db.connection.commit()
                    tvEstudiantes.delete(id)
                    lblMensaje.config(text="Se ha eliminado el registro correctamente", fg="green")
                    limpiar()
                    llenar_tabla()  # Actualizar tabla después de eliminar
                except pymysql.MySQLError as e:
                    lblMensaje.config(text=f"Error al eliminar el registro: {e}", fg="red")
            else:
                lblMensaje.config(text="Seleccione un registro para eliminar", fg="red")

    def nuevo():
        if not modificar:
            if validar():
                # Obtener el siguiente ID auto-incremental
                sql_get_id = "SELECT IFNULL(MAX(id), 0) + 1 FROM tipodecuenta"
                db.cursor.execute(sql_get_id)
                nuevo_id = db.cursor.fetchone()[0]
                
                sql = "INSERT INTO tipodecuenta (id, nombreTipoCuenta) VALUES (%s, %s)"
                try:
                    db.cursor.execute(sql, (nuevo_id, nombre_tipo_cuenta.get()))
                    db.connection.commit()
                    lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                    llenar_tabla()
                    limpiar()
                except pymysql.MySQLError as e:
                    lblMensaje.config(text=f"Error al guardar el registro: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarFalse()

    def actualizar():
        if modificar:
            if validar():
                id = id_tipo_cuenta.get()  # Tomar el ID del campo de texto
                if id:
                    sql = "UPDATE tipodecuenta SET nombreTipoCuenta=%s WHERE id=%s"
                    try:
                        db.cursor.execute(sql, (nombre_tipo_cuenta.get(), id))
                        db.connection.commit()
                        lblMensaje.config(text="Se ha guardado el registro con éxito", fg="green")
                        llenar_tabla()
                        limpiar()
                    except pymysql.MySQLError as e:
                        lblMensaje.config(text=f"Error al actualizar el registro: {e}", fg="red")
            else:
                lblMensaje.config(text="Los campos no deben estar vacíos", fg="red")
        else:
            modificarTrue()

    def vaciar_tabla():
        filas = tvEstudiantes.get_children()
        for fila in filas:
            tvEstudiantes.delete(fila)

    def llenar_tabla():
        vaciar_tabla()
        global total_records
        offset = (page_number - 1) * records_per_page
        sql_count = "SELECT COUNT(*) FROM tipodecuenta WHERE nombreTipoCuenta LIKE %s"
        try:
            db.cursor.execute(sql_count, ('%' + filtro_nombre.get() + '%',))
            total_records = db.cursor.fetchone()[0]
            
            sql = "SELECT id, nombreTipoCuenta FROM tipodecuenta WHERE nombreTipoCuenta LIKE %s LIMIT %s OFFSET %s"
            db.cursor.execute(sql, ('%' + filtro_nombre.get() + '%', records_per_page, offset))
            filas = db.cursor.fetchall()
            for fila in filas:
                id = fila[0]
                tvEstudiantes.insert("", 'end', id, text=id, values=(fila[0], fila[1]))
        except pymysql.MySQLError as e:
            lblMensaje.config(text=f"Error al cargar datos: {e}", fg="red")
        # Actualizar el estado de los botones de navegación
        btnPrev.config(state='normal' if page_number > 1 else 'disabled')
        btnNext.config(state='normal' if page_number * records_per_page < total_records else 'disabled')

    def limpiar():
        id_tipo_cuenta.set("")
        nombre_tipo_cuenta.set("")

    def validar():
        return len(nombre_tipo_cuenta.get()) > 0

    def modificarFalse():
        nonlocal modificar
        modificar = False
        tvEstudiantes.config(selectmode='none')
        btnNuevo.config(text="Guardar")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='disabled')

    def modificarTrue():
        nonlocal modificar
        modificar = True
        tvEstudiantes.config(selectmode='browse')
        btnNuevo.config(text="Nuevo")
        btnModificar.config(text="Modificar")
        btnEliminar.config(state='normal')

    def paginacion():
        llenar_tabla()
        # Actualizar el estado de los botones de navegación
        btnPrev.config(state='normal' if page_number > 1 else 'disabled')
        btnNext.config(state='normal' if page_number * records_per_page < total_records else 'disabled')

    def prev_page():
        global page_number
        if page_number > 1:
            page_number -= 1
            paginacion()

    def next_page():
        global page_number
        if page_number * records_per_page < total_records:
            page_number += 1
            paginacion()

    # Crear el marco de la ventana
    marco = tk.LabelFrame(ventana, text="Formulario Tipo de Cuenta")
    marco.place(x=50, y=50, width=1500, height=1400)

    # Botones para cerrar y minimizar
    btnCerrar = tk.Button(marco, text="X", command=ventana.destroy, bg="red", fg="white")
    btnCerrar.place(x=470, y=10, width=20, height=20)
    
    btnMinimizar = tk.Button(marco, text="-", command=lambda: ventana.iconify(), bg="yellow", fg="black")
    btnMinimizar.place(x=450, y=10, width=20, height=20)

    tk.Label(marco, text="ID").grid(column=0, row=0, padx=5, pady=5)
    txtIdTipoCuenta = tk.Entry(marco, textvariable=id_tipo_cuenta, state='disabled')
    txtIdTipoCuenta.grid(column=1, row=0)

    tk.Label(marco, text="Nombre").grid(column=0, row=1, padx=5, pady=5)
    txtNombreTipoCuenta = tk.Entry(marco, textvariable=nombre_tipo_cuenta)
    txtNombreTipoCuenta.grid(column=1, row=1)

    tk.Label(marco, text="Filtro Nombre").grid(column=0, row=2, padx=5, pady=5)
    txtFiltroNombre = tk.Entry(marco, textvariable=filtro_nombre)
    txtFiltroNombre.grid(column=1, row=2)
    filtro_nombre.trace_add('write', aplicar_filtro)

    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.grid(column=0, row=3, columnspan=4)

    tvEstudiantes = ttk.Treeview(marco, selectmode='none')  # Desactivar selección
    tvEstudiantes["columns"] = ("ID", "Nombre")
    tvEstudiantes.column("#0", width=0, stretch='no')
    tvEstudiantes.column("ID", width=150, anchor='center')
    tvEstudiantes.column("Nombre", width=150, anchor='center')
    tvEstudiantes.heading("#0", text="")
    tvEstudiantes.heading("ID", text="ID", anchor='center')
    tvEstudiantes.heading("Nombre", text="Nombre", anchor='center')
    tvEstudiantes.grid(column=0, row=4, columnspan=4, padx=5)
    tvEstudiantes.bind("<<TreeviewSelect>>", seleccionar)

    # Botones de acción
    btnEliminar = tk.Button(marco, text="Eliminar", command=lambda: eliminar())
    btnEliminar.grid(column=1, row=6, padx=5, pady=10)
    
    btnNuevo = tk.Button(marco, text="Guardar", command=lambda: nuevo())
    btnNuevo.grid(column=2, row=6, padx=5, pady=10)
    
    btnModificar = tk.Button(marco, text="Modificar", command=lambda: actualizar())
    btnModificar.grid(column=3, row=6, padx=5, pady=10)

    # Botones de navegación
    btnPrev = tk.Button(marco, text="<< Anterior", command=prev_page)
    btnPrev.grid(column=0, row=7, pady=10, sticky='w')
    
    btnNext = tk.Button(marco, text="Siguiente >>", command=next_page)
    btnNext.grid(column=3, row=7, pady=10, sticky='e')

    llenar_tabla()


class DataBase:
    def __init__(self):
        self.connection = pymysql.connect(
            host='localhost',
            user='root',
            password='',
            db='contrataciones'
        )
        self.cursor = self.connection.cursor()

class DataBase:
    def __init__(self):
        self.connection = pymysql.connect(
            host='localhost',
            user='root',
            password='',
            db='contrataciones'
        )
        self.cursor = self.connection.cursor()

def mostrar_contrato(ventana):
    db = DataBase()
    modificar = False
    descripcion_contrato = tk.StringVar()
    autorizacion_contratos = tk.StringVar()
    valor_contrato = tk.StringVar()

    # Variables de paginación
    page_number = 1
    records_per_page = 10

    # Variables de filtro
    filtro_cliente = tk.StringVar()
    filtro_tipo_contrato = tk.StringVar()
    filtro_cargo = tk.StringVar()
    filtro_dependencia = tk.StringVar()
    filtro_autorizacion = tk.StringVar()
    filtro_valor = tk.StringVar()

    def cargar_combo_data():
        try:
            # Tipo de Contrato
            sql = "SELECT nombreTipoContrato FROM tipodecontrato"
            db.cursor.execute(sql)
            tipodecontrato = db.cursor.fetchall()
            comboTipoContrato['values'] = [row[0] for row in tipodecontrato]
            comboFiltroTipoContrato['values'] = [row[0] for row in tipodecontrato]

            # Cargo
            sql = "SELECT nombreCargo FROM cargo"
            db.cursor.execute(sql)
            cargos = db.cursor.fetchall()
            comboCargo['values'] = [row[0] for row in cargos]
            comboFiltroCargo['values'] = [row[0] for row in cargos]

            # Dependencia
            sql = "SELECT nombreDependencia FROM dependencia"
            db.cursor.execute(sql)
            dependencias = db.cursor.fetchall()
            comboDependencia['values'] = [row[0] for row in dependencias]
            comboFiltroDependencia['values'] = [row[0] for row in dependencias]

            # Clientes - Solo mostrar clientes activos
            sql = """
            SELECT CONCAT_WS(' ', primerNombre, segundoNombre, primerApellido, segundoApellido) AS nombreCompleto 
            FROM clientes
            JOIN estado ON clientes.estado = estado.id
            WHERE estado.tipoEstado = 'activo'
            """
            db.cursor.execute(sql)
            clientes = db.cursor.fetchall()
            comboClientes['values'] = [row[0] for row in clientes]
            comboFiltroCliente['values'] = [row[0] for row in clientes]

        except Exception as e:
            print(f"Error al cargar datos en combobox: {e}")

    def limpiar_campos():
        descripcion_contrato.set("")
        autorizacion_contratos.set("")
        valor_contrato.set("")
        comboClientes.set("")
        comboTipoContrato.set("")
        comboCargo.set("")
        comboDependencia.set("")
        date_vigencia.set_date(datetime.today())
        date_terminacion.set_date(datetime.today())

    def llenar_tabla():
        for item in tabla.get_children():
            tabla.delete(item)

        try:
            offset = (page_number - 1) * records_per_page
            sql = """
            SELECT contrato.id, CONCAT_WS(' ', clientes.primerNombre, clientes.segundoNombre, clientes.primerApellido, clientes.segundoApellido),
                   tipodecontrato.nombreTipoContrato, cargo.nombreCargo, dependencia.nombreDependencia, 
                   contrato.descripcionContrato, contrato.Vigencia, contrato.terminacion, contrato.autorizacionContratos, contrato.valorContrato
            FROM contrato
            JOIN clientes ON contrato.idClientes = clientes.id
            JOIN tipodecontrato ON contrato.idTipoContrato = tipodecontrato.id
            JOIN cargo ON contrato.idCargo = cargo.id
            JOIN dependencia ON contrato.idDependecia = dependencia.id
            WHERE (%s IS NULL OR CONCAT_WS(' ', clientes.primerNombre, clientes.segundoNombre, clientes.primerApellido, clientes.segundoApellido) = %s)
              AND (%s IS NULL OR tipodecontrato.nombreTipoContrato = %s)
              AND (%s IS NULL OR cargo.nombreCargo = %s)
              AND (%s IS NULL OR dependencia.nombreDependencia = %s)
              AND (%s IS NULL OR contrato.autorizacionContratos = %s)
              AND (%s IS NULL OR contrato.valorContrato = %s)
            LIMIT %s OFFSET %s
            """
            filtros = (
                filtro_cliente.get() if filtro_cliente.get() else None,
                filtro_cliente.get() if filtro_cliente.get() else None,
                filtro_tipo_contrato.get() if filtro_tipo_contrato.get() else None,
                filtro_tipo_contrato.get() if filtro_tipo_contrato.get() else None,
                filtro_cargo.get() if filtro_cargo.get() else None,
                filtro_cargo.get() if filtro_cargo.get() else None,
                filtro_dependencia.get() if filtro_dependencia.get() else None,
                filtro_dependencia.get() if filtro_dependencia.get() else None,
                filtro_autorizacion.get() if filtro_autorizacion.get() else None,
                filtro_autorizacion.get() if filtro_autorizacion.get() else None,
                filtro_valor.get() if filtro_valor.get() else None,
                filtro_valor.get() if filtro_valor.get() else None,
                records_per_page,
                offset
            )
            db.cursor.execute(sql, filtros)
            contratos = db.cursor.fetchall()

            for contrato in contratos:
                tabla.insert('', 'end', values=contrato)

        except Exception as e:
            print(f"Error al llenar la tabla: {e}")

    def seleccionar_contrato(event):
        nonlocal modificar
        modificar = True

        selected_item = tabla.selection()
        if selected_item:
            item = tabla.item(selected_item)
            contrato = item['values']

            descripcion_contrato.set(contrato[5])
            date_vigencia.set_date(contrato[6])
            date_terminacion.set_date(contrato[7])
            autorizacion_contratos.set(contrato[8])
            valor_contrato.set(contrato[9])
            comboClientes.set(contrato[1])
            comboTipoContrato.set(contrato[2])
            comboCargo.set(contrato[3])
            comboDependencia.set(contrato[4])

    def guardar_contrato():
        if modificar:
            actualizar_contrato()
        else:
            insertar_contrato()

    def insertar_contrato():
        try:
            if not date_vigencia.get_date() or not date_terminacion.get_date():
                print("Error: Las fechas de vigencia y terminación son obligatorias.")
                return

            sql_cliente = """
            SELECT id FROM clientes 
            WHERE CONCAT_WS(' ', primerNombre, segundoNombre, primerApellido, segundoApellido) = %s
            LIMIT 1
            """
            db.cursor.execute(sql_cliente, (comboClientes.get(),))
            cliente_id = db.cursor.fetchone()

            sql_tipo_contrato = """
            SELECT id FROM tipodecontrato 
            WHERE nombreTipoContrato = %s
            LIMIT 1
            """
            db.cursor.execute(sql_tipo_contrato, (comboTipoContrato.get(),))
            tipo_contrato_id = db.cursor.fetchone()

            sql_cargo = """
            SELECT id FROM cargo 
            WHERE nombreCargo = %s
            LIMIT 1
            """
            db.cursor.execute(sql_cargo, (comboCargo.get(),))
            cargo_id = db.cursor.fetchone()

            sql_dependencia = """
            SELECT id FROM dependencia 
            WHERE nombreDependencia = %s
            LIMIT 1
            """
            db.cursor.execute(sql_dependencia, (comboDependencia.get(),))
            dependencia_id = db.cursor.fetchone()

            if cliente_id and tipo_contrato_id and cargo_id and dependencia_id:
                sql_insert = """
                INSERT INTO contrato 
                (idClientes, idTipoContrato, idCargo, idDependecia, descripcionContrato, Vigencia, terminacion, autorizacionContratos, valorContrato) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                valores = (
                    cliente_id[0],
                    tipo_contrato_id[0],
                    cargo_id[0],
                    dependencia_id[0],
                    descripcion_contrato.get(),
                    date_vigencia.get_date(),
                    date_terminacion.get_date(),
                    autorizacion_contratos.get(),
                    valor_contrato.get()
                )
                db.cursor.execute(sql_insert, valores)
                db.cursor.connection.commit()

                limpiar_campos()
                llenar_tabla()
            else:
                print("Error: No se encontraron todos los IDs requeridos.")

        except Exception as e:
            print(f"Error al insertar contrato: {e}")
            db.cursor.connection.rollback()

    def actualizar_contrato():
        try:
            selected_item = tabla.selection()
            if selected_item:
                item = tabla.item(selected_item)
                contrato_id = item['values'][0]

                sql_cliente = """
                SELECT id FROM clientes 
                WHERE CONCAT_WS(' ', primerNombre, segundoNombre, primerApellido, segundoApellido) = %s
                LIMIT 1
                """
                db.cursor.execute(sql_cliente, (comboClientes.get(),))
                cliente_id = db.cursor.fetchone()

                sql_tipo_contrato = """
                SELECT id FROM tipodecontrato 
                WHERE nombreTipoContrato = %s
                LIMIT 1
                """
                db.cursor.execute(sql_tipo_contrato, (comboTipoContrato.get(),))
                tipo_contrato_id = db.cursor.fetchone()

                sql_cargo = """
                SELECT id FROM cargo 
                WHERE nombreCargo = %s
                LIMIT 1
                """
                db.cursor.execute(sql_cargo, (comboCargo.get(),))
                cargo_id = db.cursor.fetchone()

                sql_dependencia = """
                SELECT id FROM dependencia 
                WHERE nombreDependencia = %s
                LIMIT 1
                """
                db.cursor.execute(sql_dependencia, (comboDependencia.get(),))
                dependencia_id = db.cursor.fetchone()

                if cliente_id and tipo_contrato_id and cargo_id and dependencia_id:
                    sql_update = """
                    UPDATE contrato 
                    SET idClientes = %s,
                        idTipoContrato = %s,
                        idCargo = %s,
                        idDependecia = %s,
                        descripcionContrato = %s, 
                        Vigencia = %s,
                        terminacion = %s,
                        autorizacionContratos = %s, 
                        valorContrato = %s
                    WHERE id = %s
                    """
                    valores = (
                        cliente_id[0],
                        tipo_contrato_id[0],
                        cargo_id[0],
                        dependencia_id[0],
                        descripcion_contrato.get(),
                        date_vigencia.get_date(),
                        date_terminacion.get_date(),
                        autorizacion_contratos.get(),
                        valor_contrato.get(),
                        contrato_id
                    )
                    db.cursor.execute(sql_update, valores)
                    db.cursor.connection.commit()

                    limpiar_campos()
                    llenar_tabla()
                else:
                    print("Error: No se encontraron todos los IDs requeridos.")

        except Exception as e:
            print(f"Error al actualizar contrato: {e}")
            db.cursor.connection.rollback()

    def eliminar_contrato():
        try:
            selected_item = tabla.selection()
            if selected_item:
                item = tabla.item(selected_item)
                contrato_id = item['values'][0]

                sql = "DELETE FROM contrato WHERE id = %s"
                db.cursor.execute(sql, (contrato_id,))
                db.cursor.connection.commit()

                limpiar_campos()
                llenar_tabla()

        except Exception as e:
            print(f"Error al eliminar contrato: {e}")
            db.cursor.connection.rollback()

    def ir_a_pagina_anterior():
        nonlocal page_number
        if page_number > 1:
            page_number -= 1
            llenar_tabla()
            actualizar_botones_paginacion()

    def ir_a_pagina_siguiente():
        nonlocal page_number
        page_number += 1
        llenar_tabla()
        actualizar_botones_paginacion()

    def mostrar_todos_los_registros():
        filtro_cliente.set("")
        filtro_tipo_contrato.set("")
        filtro_cargo.set("")
        filtro_dependencia.set("")
        filtro_autorizacion.set("")
        filtro_valor.set("")
        llenar_tabla()

    def actualizar_botones_paginacion():
        btnAnterior.config(state=tk.NORMAL if page_number > 1 else tk.DISABLED)

    # Crear ventana
    marco = tk.Frame(ventana)
    marco.pack(fill='both', expand=True)

    # Botón de cerrar ventana
    btnCerrar = tk.Button(marco, text="X", command=ventana.destroy, bg="red", fg="white")
    btnCerrar.pack(anchor="ne", padx=5, pady=5)

    # Botón de minimizar
    btnMinimizar = tk.Button(marco, text="-", command=lambda: ventana.iconify(), bg="yellow", fg="black")
    btnMinimizar.pack(anchor="ne", padx=5, pady=5)

    # Contenedor para los formularios y la tabla
    contenedor_formulario = tk.Frame(marco)
    contenedor_formulario.pack(fill='x', padx=10, pady=10)

    # Filas para organizar los inputs y selects
    for i in range(3):
        contenedor_fila = tk.Frame(contenedor_formulario)
        contenedor_fila.pack(fill='x', pady=5)

        if i == 0:
            tk.Label(contenedor_fila, text="Cliente:").grid(row=0, column=0, sticky="e")
            comboClientes = ttk.Combobox(contenedor_fila)
            comboClientes.grid(row=0, column=1, sticky="ew")

            tk.Label(contenedor_fila, text="Tipo de Contrato:").grid(row=0, column=2, sticky="e")
            comboTipoContrato = ttk.Combobox(contenedor_fila)
            comboTipoContrato.grid(row=0, column=3, sticky="ew")

            tk.Label(contenedor_fila, text="Cargo:").grid(row=0, column=4, sticky="e")
            comboCargo = ttk.Combobox(contenedor_fila)
            comboCargo.grid(row=0, column=5, sticky="ew")

        elif i == 1:
            tk.Label(contenedor_fila, text="Dependencia:").grid(row=0, column=0, sticky="e")
            comboDependencia = ttk.Combobox(contenedor_fila)
            comboDependencia.grid(row=0, column=1, sticky="ew")

            tk.Label(contenedor_fila, text="Descripción:").grid(row=0, column=2, sticky="e")
            tk.Entry(contenedor_fila, textvariable=descripcion_contrato).grid(row=0, column=3, sticky="ew")

            tk.Label(contenedor_fila, text="Vigencia:").grid(row=0, column=4, sticky="e")
            date_vigencia = DateEntry(contenedor_fila, width=20, background='darkblue',
                            foreground='white', borderwidth=2, date_pattern='y-mm-dd')
            date_vigencia.grid(row=0, column=5, sticky="ew")

        elif i == 2:
            tk.Label(contenedor_fila, text="Terminación:").grid(row=0, column=0, sticky="e")
            date_terminacion = DateEntry(contenedor_fila, width=20, background='darkblue',
                            foreground='white', borderwidth=2, date_pattern='y-mm-dd')
            date_terminacion.grid(row=0, column=1, sticky="ew")

            tk.Label(contenedor_fila, text="Autorización:").grid(row=0, column=2, sticky="e")
            tk.Entry(contenedor_fila, textvariable=autorizacion_contratos).grid(row=0, column=3, sticky="ew")

            tk.Label(contenedor_fila, text="Valor:").grid(row=0, column=4, sticky="e")
            tk.Entry(contenedor_fila, textvariable=valor_contrato).grid(row=0, column=5, sticky="ew")

    # Botones
    btnGuardar = tk.Button(contenedor_formulario, text="Guardar", command=guardar_contrato)
    btnGuardar.pack(side='left', padx=5, pady=5)

    btnEliminar = tk.Button(contenedor_formulario, text="Eliminar", command=eliminar_contrato)
    btnEliminar.pack(side='left', padx=5, pady=5)

    btnLimpiar = tk.Button(contenedor_formulario, text="Limpiar", command=limpiar_campos)
    btnLimpiar.pack(side='left', padx=5, pady=5)

    # Tabla (Treeview)
    tabla = ttk.Treeview(marco, columns=('ID', 'Cliente', 'Tipo Contrato', 'Cargo', 'Dependencia', 'Descripción', 'Vigencia', 'Terminación', 'Autorización', 'Valor'), show='headings')
    tabla.pack(fill='both', expand=True, padx=10, pady=10)

    for col in tabla['columns']:
        tabla.heading(col, text=col)

    tabla.bind('<<TreeviewSelect>>', seleccionar_contrato)

    # Botones de Paginación
    contenedor_paginacion = tk.Frame(marco)
    contenedor_paginacion.pack(fill='x', padx=10, pady=5)

    btnAnterior = tk.Button(contenedor_paginacion, text="<< Anterior", command=ir_a_pagina_anterior)
    btnAnterior.pack(side='left', fill='x', expand=True)

    btnSiguiente = tk.Button(contenedor_paginacion, text="Siguiente >>", command=ir_a_pagina_siguiente)
    btnSiguiente.pack(side='right', fill='x', expand=True)

    # Campos de filtro
    contenedor_filtros = tk.Frame(marco)
    contenedor_filtros.pack(fill='x', padx=10, pady=5)

    tk.Label(contenedor_filtros, text="Filtro Cliente:").pack(side='left')
    comboFiltroCliente = ttk.Combobox(contenedor_filtros, textvariable=filtro_cliente)
    comboFiltroCliente.pack(side='left', fill='x', expand=True)
    filtro_cliente.trace_add("write", lambda *args: llenar_tabla())

    tk.Label(contenedor_filtros, text="Filtro Tipo de Contrato:").pack(side='left')
    comboFiltroTipoContrato = ttk.Combobox(contenedor_filtros, textvariable=filtro_tipo_contrato)
    comboFiltroTipoContrato.pack(side='left', fill='x', expand=True)
    filtro_tipo_contrato.trace_add("write", lambda *args: llenar_tabla())

    tk.Label(contenedor_filtros, text="Filtro Cargo:").pack(side='left')
    comboFiltroCargo = ttk.Combobox(contenedor_filtros, textvariable=filtro_cargo)
    comboFiltroCargo.pack(side='left', fill='x', expand=True)
    filtro_cargo.trace_add("write", lambda *args: llenar_tabla())

    tk.Label(contenedor_filtros, text="Filtro Dependencia:").pack(side='left')
    comboFiltroDependencia = ttk.Combobox(contenedor_filtros, textvariable=filtro_dependencia)
    comboFiltroDependencia.pack(side='left', fill='x', expand=True)
    filtro_dependencia.trace_add("write", lambda *args: llenar_tabla())

    tk.Label(contenedor_filtros, text="Filtro Autorización:").pack(side='left')
    tk.Entry(contenedor_filtros, textvariable=filtro_autorizacion).pack(side='left', fill='x', expand=True)
    filtro_autorizacion.trace_add("write", lambda *args: llenar_tabla())

    tk.Label(contenedor_filtros, text="Filtro Valor:").pack(side='left')
    tk.Entry(contenedor_filtros, textvariable=filtro_valor).pack(side='left', fill='x', expand=True)
    filtro_valor.trace_add("write", lambda *args: llenar_tabla())

    # Botón de Mostrar Todos los Registros
    btnMostrarTodos = tk.Button(marco, text="Mostrar Todos los Registros", command=mostrar_todos_los_registros)
    btnMostrarTodos.pack(fill='x', padx=10, pady=5)

    llenar_tabla()
    cargar_combo_data()
    actualizar_botones_paginacion()

class DataBase:
    def __init__(self):
        self.connection = pymysql.connect(
            host='localhost',
            user='root',
            password='',
            db='contrataciones'
        )
        self.cursor = self.connection.cursor()

def mostrar_usuario(ventana):
    db = DataBase()
    global page_number, records_per_page, total_records
    page_number = 1
    records_per_page = 10
    total_records = 0

    modificar = False
    numero_cedula_cliente = tk.StringVar()
    correo_cli = tk.StringVar()
    contrasena_cli = tk.StringVar()
    
    filtro_cedula = tk.StringVar()
    filtro_correo = tk.StringVar()
    
    lblMensaje = tk.Label()  # Label para mostrar mensajes

    def limpiar_campos():
        numero_cedula_cliente.set("")
        correo_cli.set("")
        contrasena_cli.set("")
        filtro_cedula.set("")
        filtro_correo.set("")

    def llenar_tabla():
        # Limpiar la tabla
        for item in tabla.get_children():
            tabla.delete(item)

        # Cargar datos desde la base de datos
        offset = (page_number - 1) * records_per_page
        try:
            sql_count = "SELECT COUNT(*) FROM usuarios"
            db.cursor.execute(sql_count)
            global total_records
            total_records = db.cursor.fetchone()[0]

            filtro_cedula_valor = filtro_cedula.get().strip()
            filtro_correo_valor = filtro_correo.get().strip()

            sql = """
            SELECT id, numeroCedulaCliente, correoCli, contraseñaCli
            FROM usuarios
            WHERE (%s = '' OR numeroCedulaCliente LIKE %s)
              AND (%s = '' OR correoCli LIKE %s)
            LIMIT %s OFFSET %s
            """
            db.cursor.execute(sql, (filtro_cedula_valor, f"%{filtro_cedula_valor}%",
                                     filtro_correo_valor, f"%{filtro_correo_valor}%",
                                     records_per_page, offset))
            usuarios = db.cursor.fetchall()

            for usuario in usuarios:
                tabla.insert('', 'end', values=usuario)

        except Exception as e:
            print(f"Error al llenar la tabla: {e}")

        # Actualizar el estado de los botones de navegación
        btnPrev.config(state='normal' if page_number > 1 else 'disabled')
        btnNext.config(state='normal' if page_number * records_per_page < total_records else 'disabled')

    def seleccionar_usuario(event):
        nonlocal modificar
        modificar = True

        selected_item = tabla.selection()
        if selected_item:
            item = tabla.item(selected_item)
            usuario = item['values']

            # Llenar los campos con los valores seleccionados
            numero_cedula_cliente.set(usuario[1])
            correo_cli.set(usuario[2])
            contrasena_cli.set(usuario[3])

    def guardar_usuario():
        if not validar_campos_completos():
            lblMensaje.config(text="Por favor, complete todos los campos antes de guardar.", fg="red")
            return

        if modificar:
            actualizar_usuario()
        else:
            insertar_usuario()

    def validar_campos_completos():
        return all([
            numero_cedula_cliente.get().strip(),
            correo_cli.get().strip(),
            contrasena_cli.get().strip()
        ])

    def insertar_usuario():
        try:
            sql = """
            INSERT INTO usuarios (numeroCedulaCliente, correoCli, contraseñaCli)
            VALUES (%s, %s, %s)
            """
            valores = (
                numero_cedula_cliente.get(),
                correo_cli.get(),
                contrasena_cli.get()
            )
            db.cursor.execute(sql, valores)
            db.cursor.connection.commit()

            limpiar_campos()
            llenar_tabla()
            lblMensaje.config(text="Se ha guardado el registro con éxito.", fg="green")

        except Exception as e:
            print(f"Error al insertar usuario: {e}")
            db.cursor.connection.rollback()
            lblMensaje.config(text="Error al guardar el registro.", fg="red")

    def actualizar_usuario():
        try:
            selected_item = tabla.selection()
            if selected_item:
                item = tabla.item(selected_item)
                usuario_id = item['values'][0]

                sql = """
                UPDATE usuarios 
                SET numeroCedulaCliente = %s, 
                    correoCli = %s, 
                    contraseñaCli = %s
                WHERE id = %s
                """
                valores = (
                    numero_cedula_cliente.get(),
                    correo_cli.get(),
                    contrasena_cli.get(),
                    usuario_id
                )
                db.cursor.execute(sql, valores)
                db.cursor.connection.commit()

                limpiar_campos()
                llenar_tabla()
                lblMensaje.config(text="Se ha actualizado el registro con éxito.", fg="green")

        except Exception as e:
            print(f"Error al actualizar usuario: {e}")
            db.cursor.connection.rollback()
            lblMensaje.config(text="Error al actualizar el registro.", fg="red")

    def eliminar_usuario():
        try:
            selected_item = tabla.selection()
            if selected_item:
                item = tabla.item(selected_item)
                usuario_id = item['values'][0]

                sql = "DELETE FROM usuarios WHERE id = %s"
                db.cursor.execute(sql, (usuario_id,))
                db.cursor.connection.commit()

                limpiar_campos()
                llenar_tabla()
                lblMensaje.config(text="Se ha eliminado el registro correctamente.", fg="green")

        except Exception as e:
            print(f"Error al eliminar usuario: {e}")
            db.cursor.connection.rollback()
            lblMensaje.config(text="Error al eliminar el registro.", fg="red")

    def paginacion():
        llenar_tabla()
        # Actualizar el estado de los botones de navegación
        btnPrev.config(state='normal' if page_number > 1 else 'disabled')
        btnNext.config(state='normal' if page_number * records_per_page < total_records else 'disabled')

    def prev_page():
        global page_number
        if page_number > 1:
            page_number -= 1
            paginacion()

    def next_page():
        global page_number
        if page_number * records_per_page < total_records:
            page_number += 1
            paginacion()

    def aplicar_filtros(event=None):
        # Llenar la tabla con los filtros aplicados
        llenar_tabla()

    def mostrar_todos():
        # Limpiar filtros
        filtro_cedula.set("")
        filtro_correo.set("")
        aplicar_filtros()

    marco = tk.LabelFrame(ventana, text="Formulario Usuario")
    marco.place(x=50, y=50, width=1800, height=1600)

    # Botón de cerrar
    btnCerrar = tk.Button(marco, text="X", command=ventana.destroy, bg="red", fg="white")
    btnCerrar.place(x=770, y=10, width=20, height=20)

    # Botón de minimizar
    btnMinimizar = tk.Button(marco, text="-", command=lambda: ventana.iconify(), bg="yellow", fg="black")
    btnMinimizar.place(x=740, y=10, width=20, height=20)

    # Etiquetas y entradas para registrar
    tk.Label(marco, text="Cédula del Cliente:").place(x=50, y=50)
    tk.Entry(marco, textvariable=numero_cedula_cliente).place(x=200, y=50)

    tk.Label(marco, text="Correo:").place(x=50, y=100)
    tk.Entry(marco, textvariable=correo_cli).place(x=200, y=100)

    tk.Label(marco, text="Contraseña:").place(x=50, y=150)
    tk.Entry(marco, textvariable=contrasena_cli, show="*").place(x=200, y=150)

    # Botones para registrar
    btnGuardar = tk.Button(marco, text="Guardar", command=guardar_usuario)
    btnGuardar.place(x=200, y=200)

    btnEliminar = tk.Button(marco, text="Eliminar", command=eliminar_usuario)
    btnEliminar.place(x=300, y=200)

    btnLimpiar = tk.Button(marco, text="Limpiar", command=limpiar_campos)
    btnLimpiar.place(x=400, y=200)

    # Label para mensajes
    lblMensaje = tk.Label(marco, text="Aquí van los mensajes", fg="green")
    lblMensaje.place(x=200, y=240)  # Ajusta la posición según sea necesario

    # Tabla (Treeview)
    columnas = ('ID', 'Cédula Cliente', 'Correo', 'Contraseña')
    tabla = ttk.Treeview(marco, columns=columnas, show='headings')
    tabla.place(x=50, y=280, width=700, height=300)

    for col in columnas:
        tabla.heading(col, text=col)

    tabla.bind('<<TreeviewSelect>>', seleccionar_usuario)

    # Botones de navegación
    btnPrev = tk.Button(marco, text="<< Anterior", command=prev_page)
    btnPrev.place(x=50, y=590, width=80, height=30)
    
    btnNext = tk.Button(marco, text="Siguiente >>", command=next_page)
    btnNext.place(x=670, y=590, width=80, height=30)

    # Etiquetas y entradas para filtros
    tk.Label(marco, text="Filtrar por Cédula del Cliente:").place(x=50, y=350)
    filtro_cedula_entry = tk.Entry(marco, textvariable=filtro_cedula)
    filtro_cedula_entry.place(x=200, y=350)
    filtro_cedula_entry.bind("<KeyRelease>", aplicar_filtros)

    tk.Label(marco, text="Filtrar por Correo:").place(x=50, y=400)
    filtro_correo_entry = tk.Entry(marco, textvariable=filtro_correo)
    filtro_correo_entry.place(x=200, y=400)
    filtro_correo_entry.bind("<KeyRelease>", aplicar_filtros)

    # Botón para mostrar todos los registros
    btnMostrarTodos = tk.Button(marco, text="Mostrar Todos los Registros", command=mostrar_todos)
    btnMostrarTodos.place(x=50, y=450, width=200, height=30)

    # Llenar la tabla inicialmente
    paginacion()

def cerrar_sesion():
    global usuario_iniciado
    usuario_iniciado = False
    pantalla1.destroy()  # Cierra la ventana de bienvenida y regresa a la pantalla principal

menu_pantalla()
